"""Campaign template file generation for Campaign Optimizer."""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Optional, Dict, Any


class CampaignTemplateGenerator:
    """Generates template Excel files for campaign creation."""

    def __init__(self):
        """Initialize campaign template generator."""
        self.sheet_name = "Campaign Configuration"
        self.columns = [
            "My ASIN",
            "Product Type",
            "Niche",
            "Testing Bid",
            "Testing PT Bid",
            "Phrase Bid",
            "Broad Bid",
            "Expanded Bid",
            "Halloween Testing Bid",
            "Halloween Testing Bid",
            "Halloween Testing PT Bid",
            "Halloween Phrase Bid",
            "Halloween Broad Bid",
            "Halloween Expanded Bid"
        ]

    def generate_template(self) -> bytes:
        """
        Generate a campaign template Excel file.
        
        Returns:
            bytes: Excel file as bytes
        """
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Campaign Configuration sheet
        ws = wb.create_sheet(self.sheet_name)
        
        # Add headers
        for col_idx, header in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # Center alignment for headers
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set uniform column widths
        for col_idx in range(1, 15):  # 14 columns (הורדנו את Hero Keywords)
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
            ws.column_dimensions[col_letter].width = 20
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _set_column_widths(self, worksheet):
        """Set appropriate column widths for the template."""
        # All columns with uniform width
        for col_idx in range(1, 15):  # 14 columns (הורדנו את Hero Keywords)
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
            worksheet.column_dimensions[col_letter].width = 20
    
    def validate_template_structure(self, file_data: bytes) -> tuple[bool, str]:
        """
        Validate that uploaded file has correct campaign template structure.
        
        Args:
            file_data: Excel file as bytes
            
        Returns:
            Tuple of (is_valid, message)
        """
        try:
            # Read Excel file
            excel_data = pd.read_excel(file_data, sheet_name=None)
            
            # Check if Campaign Configuration sheet exists
            if self.sheet_name not in excel_data:
                return False, f"Missing required sheet: {self.sheet_name}"
            
            # Get the sheet
            df = excel_data[self.sheet_name]
            
            # Check columns
            expected_cols = set(self.columns)
            actual_cols = set(df.columns.tolist())
            
            if expected_cols != actual_cols:
                missing = expected_cols - actual_cols
                extra = actual_cols - expected_cols
                
                issues = []
                if missing:
                    issues.append(f"Missing columns: {', '.join(missing)}")
                if extra:
                    issues.append(f"Extra columns: {', '.join(extra)}")
                
                return False, "; ".join(issues)
            
            return True, "Campaign template structure is valid"
            
        except Exception as e:
            return False, f"Error reading template: {str(e)}"
    
    def get_template_info(self) -> Dict[str, Any]:
        """
        Get information about the template structure.
        
        Returns:
            Dictionary with template information
        """
        return {
            "sheet_name": self.sheet_name,
            "columns": self.columns,
            "column_count": len(self.columns),
            "bid_columns": [col for col in self.columns if "Bid" in col],
        }