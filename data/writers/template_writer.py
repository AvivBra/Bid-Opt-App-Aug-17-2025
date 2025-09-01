"""Template file writer for portfolio optimizations."""

import pandas as pd
import io
from typing import Dict, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class TemplateWriter:
    """Writes template files for portfolio optimizations."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_top_asins_template(self) -> bytes:
        """
        Create Top ASINs template Excel file.
        
        Returns:
            Excel file as bytes
        """
        self.logger.info("Creating Top ASINs template")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Set header
        ws['A1'] = 'Top ASINs'
        
        # Style header
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        
        # Set column width
        ws.column_dimensions['A'].width = 15
        
        # Add example ASINs (optional, for user guidance)
        example_asins = [
            'B08N5WRWNW',
            'B07QRXF3QV', 
            'B09LQRS8TN'
        ]
        
        for i, asin in enumerate(example_asins, start=2):
            ws[f'A{i}'] = asin
            # Make example ASINs lighter colored to indicate they're examples
            ws[f'A{i}'].font = Font(color='888888', italic=True)
        
        # Add instruction comment (optional)
        comment_text = "Enter your top ASINs here, one per row. Delete the example ASINs above."
        ws[f'A{len(example_asins) + 3}'] = comment_text
        ws[f'A{len(example_asins) + 3}'].font = Font(size=10, italic=True, color='666666')
        
        # Convert to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        excel_bytes = buffer.getvalue()
        self.logger.info(f"Top ASINs template created: {len(excel_bytes)} bytes")
        
        return excel_bytes
    
    def create_filled_template_example(self, asins: list) -> bytes:
        """
        Create filled template example for testing.
        
        Args:
            asins: List of ASINs to include
            
        Returns:
            Excel file as bytes
        """
        self.logger.info(f"Creating filled template example with {len(asins)} ASINs")
        
        # Create DataFrame
        template_df = pd.DataFrame({'Top ASINs': asins})
        
        # Convert to Excel bytes
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            template_df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Access workbook to style header
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Style header
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            worksheet['A1'].font = header_font
            worksheet['A1'].fill = header_fill
            
            # Set column width
            worksheet.column_dimensions['A'].width = 15
        
        buffer.seek(0)
        excel_bytes = buffer.getvalue()
        
        self.logger.info(f"Filled template example created: {len(excel_bytes)} bytes")
        return excel_bytes
    
    def write_top_sheet_to_excel(self, workbook, top_asins_df: pd.DataFrame) -> None:
        """
        Add Top sheet to existing Excel workbook.
        
        Args:
            workbook: openpyxl Workbook object
            top_asins_df: DataFrame with Top ASINs data
        """
        self.logger.info("Adding Top sheet to workbook")
        
        # Create Top sheet
        top_sheet = workbook.create_sheet(title="Top")
        
        # Write header
        top_sheet['A1'] = 'Top ASINs'
        
        # Style header
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        top_sheet['A1'].font = header_font
        top_sheet['A1'].fill = header_fill
        top_sheet['A1'].alignment = header_alignment
        
        # Write ASIN data
        for i, asin in enumerate(top_asins_df['Top ASINs'], start=2):
            top_sheet[f'A{i}'] = asin
        
        # Set column width
        top_sheet.column_dimensions['A'].width = 15
        
        self.logger.info(f"Top sheet added with {len(top_asins_df)} ASINs")