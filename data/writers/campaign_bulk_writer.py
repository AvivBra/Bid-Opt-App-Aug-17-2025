"""Campaign bulk file writer."""

import pandas as pd
from io import BytesIO
from typing import Dict, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows


class CampaignBulkWriter:
    """Writes campaign data to Excel bulk upload file."""
    
    def __init__(self):
        """Initialize campaign bulk writer."""
        self.logger = logging.getLogger(__name__)
        
        # Excel formatting
        self.header_font = Font(bold=True)
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def write(
        self,
        sheets_data: Dict[str, pd.DataFrame],
        filename: Optional[str] = None
    ) -> BytesIO:
        """Write campaign data to Excel file.
        
        Args:
            sheets_data: Dictionary mapping sheet names to DataFrames
            filename: Optional filename for logging
            
        Returns:
            BytesIO object containing the Excel file
        """
        output = BytesIO()
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Expected sheet order
            sheet_order = ["Campaign", "Ad Group", "Product Ad", "Keyword"]
            
            # Add sheets in correct order
            sheets_added = 0
            for sheet_name in sheet_order:
                if sheet_name in sheets_data:
                    df = sheets_data[sheet_name]
                    
                    if df is None:
                        self.logger.warning(f"Skipping {sheet_name}: None value")
                        continue
                    
                    if df.empty:
                        self.logger.warning(f"Skipping {sheet_name}: Empty DataFrame")
                        # Still create empty sheet with headers
                        ws = wb.create_sheet(title=sheet_name)
                        self._write_empty_sheet_with_headers(ws)
                        sheets_added += 1
                    else:
                        ws = wb.create_sheet(title=sheet_name)
                        self._write_dataframe_to_sheet(ws, df)
                        self._format_worksheet(ws, sheet_name)
                        sheets_added += 1
                        
                        self.logger.info(f"Added {sheet_name} sheet with {len(df)} rows")
            
            # Add any additional sheets not in standard order
            for sheet_name, df in sheets_data.items():
                if sheet_name not in sheet_order:
                    if df is not None and not df.empty:
                        ws = wb.create_sheet(title=self._clean_sheet_name(sheet_name))
                        self._write_dataframe_to_sheet(ws, df)
                        self._format_worksheet(ws, sheet_name)
                        sheets_added += 1
                        
                        self.logger.info(f"Added additional sheet {sheet_name} with {len(df)} rows")
            
            if sheets_added == 0:
                self.logger.error("No sheets were added to the workbook")
                # Create at least one empty sheet
                ws = wb.create_sheet(title="Campaign")
                self._write_empty_sheet_with_headers(ws)
            
            # Save workbook to BytesIO
            wb.save(output)
            output.seek(0)
            
            self.logger.info(f"Successfully created Excel file with {sheets_added} sheets")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
            raise
        
        return output
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """Write DataFrame to worksheet.
        
        Args:
            ws: Worksheet object
            df: DataFrame to write
        """
        # Write data using dataframe_to_rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format headers
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.alignment = self.center_alignment
                else:
                    # Center align all data cells
                    cell.alignment = self.center_alignment
                
                # Convert None or NaN to empty string
                if pd.isna(value):
                    cell.value = ""
                # Format Campaign ID and Ad Group ID as text
                elif c_idx in [4, 5] and r_idx > 1:  # Campaign ID and Ad Group ID columns
                    cell.value = str(value) if value else ""
                    cell.number_format = '@'  # Text format
    
    def _write_empty_sheet_with_headers(self, ws):
        """Write empty sheet with standard headers.
        
        Args:
            ws: Worksheet object
        """
        headers = [
            "Product", "Entity", "Operation", "Campaign ID", "Ad Group ID",
            "Portfolio ID", "Ad ID", "Keyword ID", "Product Targeting ID",
            "Campaign Name", "Ad Group Name", "Start Date", "End Date",
            "Targeting Type", "State", "Daily Budget", "SKU", "ASIN",
            "Ad Group Default Bid", "Bid", "Keyword Text", "Native Language Keyword",
            "Native Language Locale", "Match Type", "Bidding Strategy",
            "Placement", "Percentage", "Product Targeting Expression"
        ]
        
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = self.header_font
            cell.alignment = self.center_alignment
    
    def _format_worksheet(self, ws, sheet_name: str):
        """Format worksheet with proper column widths and alignment.
        
        Args:
            ws: Worksheet object
            sheet_name: Name of the sheet
        """
        # Set column widths based on content type
        column_widths = {
            'A': 20,  # Product
            'B': 15,  # Entity
            'C': 12,  # Operation
            'D': 50,  # Campaign ID
            'E': 25,  # Ad Group ID
            'F': 15,  # Portfolio ID
            'G': 15,  # Ad ID
            'H': 15,  # Keyword ID
            'I': 25,  # Product Targeting ID
            'J': 50,  # Campaign Name
            'K': 25,  # Ad Group Name
            'L': 12,  # Start Date
            'M': 12,  # End Date
            'N': 15,  # Targeting Type
            'O': 10,  # State
            'P': 12,  # Daily Budget
            'Q': 15,  # SKU
            'R': 15,  # ASIN
            'S': 18,  # Ad Group Default Bid
            'T': 10,  # Bid
            'U': 30,  # Keyword Text
            'V': 30,  # Native Language Keyword
            'W': 20,  # Native Language Locale
            'X': 12,  # Match Type
            'Y': 25,  # Bidding Strategy
            'Z': 15,  # Placement
            'AA': 12, # Percentage
            'AB': 35  # Product Targeting Expression
        }
        
        for col_letter, width in column_widths.items():
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = width
            else:
                try:
                    ws.column_dimensions[col_letter].width = width
                except Exception:
                    pass  # Column might not exist
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
    
    def _clean_sheet_name(self, name: str) -> str:
        """Clean sheet name to meet Excel requirements.
        
        Args:
            name: Original sheet name
            
        Returns:
            Cleaned sheet name
        """
        # Remove invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '')
        
        # Truncate to 31 characters (Excel limit)
        if len(name) > 31:
            name = name[:31]
        
        return name
    
    def validate_output_file(self, output: BytesIO) -> bool:
        """Validate that the output file is valid.
        
        Args:
            output: BytesIO object with Excel file
            
        Returns:
            True if file is valid
        """
        try:
            # Check file size
            output.seek(0, 2)  # Seek to end
            file_size = output.tell()
            output.seek(0)  # Reset to beginning
            
            if file_size == 0:
                self.logger.error("Output file is empty")
                return False
            
            # Try to load with openpyxl to validate
            from openpyxl import load_workbook
            wb = load_workbook(output, read_only=True)
            sheet_count = len(wb.sheetnames)
            wb.close()
            
            if sheet_count == 0:
                self.logger.error("Output file has no sheets")
                return False
            
            output.seek(0)  # Reset for reading
            
            self.logger.info(f"Output file validated: {file_size} bytes, {sheet_count} sheets")
            return True
            
        except Exception as e:
            self.logger.error(f"Error validating output file: {str(e)}")
            return False