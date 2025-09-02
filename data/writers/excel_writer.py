"""Excel writer for output files - UPDATED for proper formatting."""

import pandas as pd
from io import BytesIO
from typing import Dict, Optional, List, Any
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ..writers.template_writer import TemplateWriter


class ExcelWriter:
    """
    Writes optimization results to Excel files with proper formatting.

    Features:
    - Creates multi-sheet Excel files
    - Applies pink highlighting to error rows
    - Formats numbers as text for long IDs
    - Centers all cell values
    - No borders on cells
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)

        # Yellow color for highlighting updated rows
        self.yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        # Blue color for ASIN PA header
        self.blue_fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )

        # Header formatting (optional - you can remove if not wanted)
        self.header_fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )

        self.header_font = Font(bold=True)
        self.blue_header_font = Font(bold=True, color="FFFFFF")  # White text on blue background

        # Center alignment for all cells
        self.center_alignment = Alignment(horizontal="center", vertical="center")

        # No borders
        self.no_border = Border()
        
        # Template writer for Top sheet handling
        self.template_writer = TemplateWriter()

    def write_excel(
        self, sheets_dict: Dict[str, pd.DataFrame], filename: Optional[str] = None
    ) -> BytesIO:
        """
        Write multiple DataFrames to an Excel file with proper formatting.

        Args:
            sheets_dict: Dictionary mapping sheet names to DataFrames
            filename: Optional filename (for logging purposes)

        Returns:
            BytesIO object containing the Excel file
        """

        output = BytesIO()

        # DEBUG: Log what ExcelWriter received
        print(f"[DEBUG ExcelWriter] write_excel() called:")
        print(f"[DEBUG ExcelWriter]   sheets_dict type: {type(sheets_dict)}")
        print(f"[DEBUG ExcelWriter]   number of sheets: {len(sheets_dict)}")
        for sheet_name, df in sheets_dict.items():
            if hasattr(df, 'shape'):
                print(f"[DEBUG ExcelWriter]   {sheet_name}: {df.shape} (DataFrame)")
            else:
                print(f"[DEBUG ExcelWriter]   {sheet_name}: {type(df)}")

        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Define expected sheet order for Portfolio Optimizer
            expected_order = ['Portfolios', 'Product Ad', 'Campaign', 'Terminal', 'Top', 'Sheet3']
            
            # Reorder sheets according to expected order, then add any remaining sheets
            ordered_sheets = []
            for sheet_name in expected_order:
                if sheet_name in sheets_dict:
                    ordered_sheets.append((sheet_name, sheets_dict[sheet_name]))
            
            # Add any remaining sheets not in expected order
            for sheet_name, df in sheets_dict.items():
                if sheet_name not in expected_order:
                    ordered_sheets.append((sheet_name, df))
            
            # Add each dataframe as a sheet in correct order
            for sheet_name, df in ordered_sheets:
                if not isinstance(df, pd.DataFrame):
                    self.logger.warning(f"Skipping {sheet_name}: not a DataFrame")
                    continue

                if df.empty and sheet_name != "Sheet3":
                    self.logger.warning(f"Skipping empty sheet: {sheet_name}")
                    continue

                # Handle special Top sheet
                if sheet_name == "Top":
                    self._add_top_sheet(wb, df)
                    continue

                # Create worksheet
                ws = wb.create_sheet(title=self._clean_sheet_name(sheet_name))

                # Remove internal columns before writing
                df_to_write = df.copy()
                internal_columns = ["_needs_highlight", "_needs_pink_highlight", "_error_type", "Camp Count", "Old Portfolio Name"]
                for col in internal_columns:
                    if col in df_to_write.columns:
                        df_to_write = df_to_write.drop(columns=[col])
                
                
                # Write dataframe to worksheet
                self._write_dataframe_to_sheet(ws, df_to_write)

                # Apply formatting
                self._format_worksheet(ws, df_to_write)

                # Highlight updated rows if present
                if "_needs_highlight" in df.columns:
                    # Handle _needs_highlight column from optimization processors
                    updated_rows = df[df["_needs_highlight"] == True].index.tolist()
                    self._highlight_updated_rows(ws, updated_rows)

            # Save workbook to BytesIO
            wb.save(output)
            output.seek(0)

            # DEBUG: Log final output details
            output_size = len(output.getvalue())
            print(f"[DEBUG ExcelWriter] Final Excel file created:")
            print(f"[DEBUG ExcelWriter]   File size: {output_size} bytes")
            print(f"[DEBUG ExcelWriter]   Sheets processed: {len(sheets_dict)}")

            self.logger.info(
                f"Successfully created Excel file with {len(sheets_dict)} sheets"
            )

        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
            raise

        return output

    def _clean_sheet_name(self, name: str) -> str:
        """
        Clean sheet name to meet Excel requirements.

        Excel sheet names must be <= 31 characters and cannot contain: : \ / ? * [ ]
        """


        # Remove invalid characters
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            name = name.replace(char, "")

        # Truncate to 31 characters
        if len(name) > 31:
            name = name[:31]


        return name

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """
        Write a DataFrame to an Excel worksheet.
        """

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            
            # Special formatting for ASIN PA column header (blue background)
            if col_name == "ASIN PA":
                cell.font = self.blue_header_font
                cell.fill = self.blue_fill
            else:
                cell.font = self.header_font
                
            cell.alignment = self.center_alignment
            cell.border = self.no_border

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Convert value to appropriate type
                col_name = df.columns[col_idx - 1]
                
                # DEBUG: Check what type of data we're getting for ID columns and problematic columns
                is_id_col_debug = any(id_keyword in col_name for id_keyword in [
                    'Product Targeting ID', 'Campaign ID', 'Ad Group ID', 'Keyword ID', 
                    'Portfolio ID', 'ASIN', 'Target ID', 'Ad ID', 'ID'
                ])
                if (is_id_col_debug or col_name in ['Budget Amount', 'Budget Start Date', 'Daily Budget', 'Start Date']) and row_idx <= 3:
                    print(f'[DEBUG excel_writer] {ws.title}.{col_name}[{row_idx-2}]: {repr(value)} (type: {type(value)}) ID_COLUMN: {is_id_col_debug}')
                
                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, (int, float)):
                    # Check if this is an ID column that should be formatted as text
                    is_id_column = any(id_keyword in col_name for id_keyword in [
                        'Product Targeting ID', 'Campaign ID', 'Ad Group ID', 'Keyword ID', 
                        'Portfolio ID', 'ASIN', 'Target ID', 'Ad ID', 'ID'
                    ])
                    
                    # Check if this is a column that should be converted to integer in the Campaign sheet
                    is_integer_column = col_name in ['Start Date', 'Daily Budget', 'End Date'] and ws.title == 'Campaign'
                    
                    # Check if this is an ID column in Product Ad sheet that should be integer
                    is_product_ad_id = (ws.title == 'Product Ad' and 
                                      col_name in ['Ad Group ID', 'Ad ID'] and
                                      isinstance(value, float) and value.is_integer())
                    
                    # Format ID columns as text to prevent scientific notation and decimal display  
                    if is_id_column:
                        # Convert to integer first to remove decimals, then to string
                        if isinstance(value, float) and value.is_integer():
                            cell.value = str(int(value))
                        else:
                            cell.value = str(value)
                        cell.number_format = "@"  # Text format
                    
                    # Convert certain Campaign columns to integers
                    elif is_integer_column:
                        if isinstance(value, float) and value.is_integer():
                            cell.value = int(value)
                        else:
                            cell.value = value
                    
                    # Convert Product Ad ID columns to integers  
                    elif is_product_ad_id:
                        cell.value = int(value)
                    
                    else:
                        cell.value = value
                        # Regular number format with 3 decimal places for bid values
                        if "bid" in col_name.lower() or col_name in [
                            "calc1",
                            "calc2",
                            "Target CPA",
                            "Base Bid",
                            "Adj. CPA",
                            "Max BA",
                            "Temp Bid",
                            "Max_Bid",
                            "calc3",
                        ]:
                            cell.number_format = "0.000"
                        # For financial columns, preserve exact values including precision artifacts
                        elif col_name in ["Spend", "Sales", "Cost", "Revenue"]:
                            # Don't apply number formatting - let the original string values be preserved
                            cell.number_format = "General"
                        else:
                            cell.number_format = "General"
                elif isinstance(value, str):
                    # Handle string values - force text format for ID columns
                    is_id_column = any(id_keyword in col_name for id_keyword in [
                        'Product Targeting ID', 'Campaign ID', 'Ad Group ID', 'Keyword ID', 
                        'Portfolio ID', 'ASIN', 'Target ID', 'Ad ID', 'ID'
                    ])
                    
                    if is_id_column:
                        # Remove .0 suffix from string ID values
                        if value.endswith('.0'):
                            clean_value = value[:-2]  # Remove '.0'
                            cell.value = clean_value
                        else:
                            cell.value = str(value)
                        cell.number_format = "@"  # Text format
                    else:
                        cell.value = str(value)
                else:
                    cell.value = str(value)

                # Apply center alignment and no border
                cell.alignment = self.center_alignment
                cell.border = self.no_border

    def _format_worksheet(self, ws, df: pd.DataFrame):
        """
        Apply formatting to the worksheet.
        """

        # Set consistent column width for ALL columns (15 characters)
        for col_idx in range(1, len(df.columns) + 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = 15

        # Ensure Old Bid column is properly positioned and formatted
        if "Old Bid" in df.columns and "Bid" in df.columns:
            old_bid_idx = df.columns.get_loc("Old Bid") + 1
            bid_idx = df.columns.get_loc("Bid") + 1

            # The processor should have already arranged columns correctly
            # Just ensure formatting is applied
            for row in range(2, ws.max_row + 1):
                old_bid_cell = ws.cell(row=row, column=old_bid_idx)
                bid_cell = ws.cell(row=row, column=bid_idx)

                # Both should have same number format
                old_bid_cell.number_format = "0.000"
                bid_cell.number_format = "0.000"

    def _highlight_updated_rows(self, ws, updated_indices: List[int]):
        """
        Highlight rows that were updated in yellow.

        Args:
            ws: Worksheet object
            updated_indices: List of row indices to highlight (0-based from DataFrame)
        """

        for idx in updated_indices:
            # Convert DataFrame index to Excel row (add 2: +1 for 0-based to 1-based, +1 for header)
            excel_row = idx + 2

            # Apply yellow fill to entire row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.fill = self.yellow_fill

    def create_working_file(
        self, optimization_results: Dict[str, Dict[str, pd.DataFrame]]
    ) -> BytesIO:
        """
        Create a Working File with all optimization results.

        Args:
            optimization_results: Nested dict of optimization name -> sheet name -> DataFrame

        Returns:
            BytesIO object containing the Working File
        """


        # Flatten the nested dictionary
        all_sheets = {}

        for opt_name, sheets in optimization_results.items():

            for sheet_name, df in sheets.items():

                # Use sheet name directly if it already includes optimization name
                if opt_name.lower() in sheet_name.lower():
                    final_sheet_name = sheet_name
                else:
                    final_sheet_name = f"{sheet_name} - {opt_name}"


                all_sheets[final_sheet_name] = df


        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"Working_File_{timestamp}.xlsx"

        self.logger.info(f"Creating Working File: {filename}")

        return self.write_excel(all_sheets, filename)

    def create_clean_file(
        self, optimization_results: Dict[str, Dict[str, pd.DataFrame]]
    ) -> BytesIO:
        """
        Create a Clean File (currently same as Working File).

        In future phases, this will remove helper columns.

        Args:
            optimization_results: Nested dict of optimization name -> sheet name -> DataFrame

        Returns:
            BytesIO object containing the Clean File
        """

        # For now, Clean File is same as Working File
        # In future, will remove helper columns here

        self.logger.info("Creating Clean File (currently same as Working File)")

        return self.create_working_file(optimization_results)
    
    def _add_top_sheet(self, workbook, top_asins_df: pd.DataFrame) -> None:
        """
        Add Top sheet to workbook using template writer.
        
        Args:
            workbook: openpyxl Workbook object
            top_asins_df: DataFrame with Top ASINs data
        """
        self.logger.info("Adding Top sheet to workbook")
        self.template_writer.write_top_sheet_to_excel(workbook, top_asins_df)
    
    def create_campaign_optimizer_1_file(self, processed_data: Dict[str, pd.DataFrame]) -> BytesIO:
        """
        Create Campaign Optimizer 1 output file with proper formatting.
        
        Args:
            processed_data: Dictionary of sheet_name -> DataFrame
            
        Returns:
            BytesIO object containing the Excel file
        """
        self.logger.info("Creating Campaign Optimizer 1 output file")
        
        # Use the standard write_excel method
        return self.write_excel(processed_data, "campaign_optimizer_1_output.xlsx")
