"""Template file generation for Bid Optimizer."""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config.constants import (
    TEMPLATE_REQUIRED_SHEETS,
    TEMPLATE_PORT_VALUES_COLUMNS,
    TEMPLATE_DELETE_FOR_60_COLUMNS,
)
from business.portfolio_optimizations.templates import TopCampaignsTemplateGenerator


class TemplateGenerator:
    """Generates template Excel files for bid optimization."""

    def __init__(self):
        self.port_values_columns = TEMPLATE_PORT_VALUES_COLUMNS
        self.delete_for_60_columns = TEMPLATE_DELETE_FOR_60_COLUMNS
        self.required_sheets = TEMPLATE_REQUIRED_SHEETS
        self.top_campaigns_generator = TopCampaignsTemplateGenerator()

    def generate_template(self) -> bytes:
        """Generate a complete template Excel file."""

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create Port Values sheet
        self._create_port_values_sheet(wb)

        # Create Top ASINs sheet
        self._create_top_asins_sheet(wb)

        # Create Delete for 60 sheet
        self._create_delete_for_60_sheet(wb)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    def _create_port_values_sheet(self, wb: Workbook):
        """Create the Port Values sheet with headers and sample data."""
        ws = wb.create_sheet("Port Values")

        # Headers
        headers = self.port_values_columns
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

    def _create_top_asins_sheet(self, wb: Workbook):
        """Create the Top ASINs sheet (currently empty - for future use)."""

        ws = wb.create_sheet("Top ASINs")

        # Header
        cell = ws.cell(row=1, column=1, value="ASIN")

        # Column width
        ws.column_dimensions["A"].width = 50

    def _create_delete_for_60_sheet(self, wb: Workbook):
        """Create the Delete for 60 sheet with headers."""

        ws = wb.create_sheet("Delete for 60")

        # Headers
        headers = self.delete_for_60_columns
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)

        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25

    def validate_template_structure(self, file_data: bytes) -> tuple[bool, str]:
        """Validate that uploaded file has correct template structure."""

        try:
            # Read Excel file
            excel_data = pd.read_excel(file_data, sheet_name=None)

            # Check required sheets exist
            missing_sheets = []
            for sheet in self.required_sheets:
                if sheet not in excel_data:
                    missing_sheets.append(sheet)

            if missing_sheets:
                return False, f"Missing sheets: {', '.join(missing_sheets)}"

            # Check Port Values columns
            port_values_df = excel_data["Port Values"]
            expected_cols = set(self.port_values_columns)
            actual_cols = set(port_values_df.columns)

            if expected_cols != actual_cols:
                missing = expected_cols - actual_cols
                extra = actual_cols - expected_cols

                issues = []
                if missing:
                    issues.append(f"Missing columns: {', '.join(missing)}")
                if extra:
                    issues.append(f"Extra columns: {', '.join(extra)}")

                return False, "; ".join(issues)

            # Check Delete for 60 columns if sheet exists
            if "Delete for 60" in excel_data:
                delete_for_60_df = excel_data["Delete for 60"]
                expected_cols = set(self.delete_for_60_columns)
                actual_cols = set(delete_for_60_df.columns)

                if expected_cols != actual_cols:
                    missing = expected_cols - actual_cols
                    extra = actual_cols - expected_cols

                    issues = []
                    if missing:
                        issues.append(
                            f"Delete for 60 - Missing columns: {', '.join(missing)}"
                        )
                    if extra:
                        issues.append(
                            f"Delete for 60 - Extra columns: {', '.join(extra)}"
                        )

                    return False, "; ".join(issues)

            return True, "Template structure is valid"

        except Exception as e:
            return False, f"Error reading template: {str(e)}"

    def get_sample_portfolios(self) -> list[str]:
        """Get list of sample portfolio names for testing."""
        return [
            "Campaign Alpha Portfolio",
            "Campaign Beta Portfolio",
            "Campaign Gamma Portfolio",
            "Test Portfolio 1",
            "Test Portfolio 2",
        ]
    
    def generate_top_asins_template(self) -> bytes:
        """Generate Top ASINs template for Organize Top Campaigns optimization."""
        return self.top_campaigns_generator.generate_template()
