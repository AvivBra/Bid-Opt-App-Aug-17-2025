#!/usr/bin/env python3
"""
Alternative Excel file loading methods
Try different libraries and approaches to load the problematic expected output file
"""

import sys
import traceback
from pathlib import Path
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def try_openpyxl_direct():
    """Try loading with openpyxl directly, bypassing pandas"""
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        expected_file = "/Applications/My Apps/Bid Opt App Aug 17, 2025/PRD/Portfilio Optimizer/Excel Examples/Output Bulk Example for all 3 opt.xlsx"
        
        logger.info("Trying openpyxl direct approach...")
        
        # Load workbook
        wb = openpyxl.load_workbook(expected_file, data_only=True)
        logger.info(f"✅ Workbook loaded successfully with openpyxl")
        logger.info(f"Sheet names: {wb.sheetnames}")
        
        # Convert to data structure
        data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []
            
            # Get all rows
            for row in ws.iter_rows(values_only=True):
                sheet_data.append(row)
            
            data[sheet_name] = sheet_data
            logger.info(f"  - {sheet_name}: {len(sheet_data)} rows")
        
        wb.close()
        return data
        
    except Exception as e:
        logger.error(f"openpyxl direct failed: {e}")
        return None

def try_xlwings():
    """Try loading with xlwings if available"""
    try:
        import xlwings as xw
        
        expected_file = "/Applications/My Apps/Bid Opt App Aug 17, 2025/PRD/Portfilio Optimizer/Excel Examples/Output Bulk Example for all 3 opt.xlsx"
        
        logger.info("Trying xlwings approach...")
        
        # Open Excel file
        wb = xw.Book(expected_file)
        logger.info(f"✅ Workbook loaded successfully with xlwings")
        
        data = {}
        for sheet in wb.sheets:
            sheet_name = sheet.name
            # Get used range
            used_range = sheet.used_range
            if used_range is not None:
                values = used_range.value
                data[sheet_name] = values
                logger.info(f"  - {sheet_name}: {len(values) if values else 0} rows")
            else:
                data[sheet_name] = []
                logger.info(f"  - {sheet_name}: empty sheet")
        
        wb.close()
        return data
        
    except ImportError:
        logger.error("xlwings not available")
        return None
    except Exception as e:
        logger.error(f"xlwings failed: {e}")
        return None

def try_pyexcel():
    """Try loading with pyexcel if available"""
    try:
        import pyexcel as pe
        
        expected_file = "/Applications/My Apps/Bid Opt App Aug 17, 2025/PRD/Portfilio Optimizer/Excel Examples/Output Bulk Example for all 3 opt.xlsx"
        
        logger.info("Trying pyexcel approach...")
        
        # Get book
        book = pe.get_book(file_name=expected_file)
        logger.info(f"✅ Workbook loaded successfully with pyexcel")
        logger.info(f"Sheet names: {list(book.sheet_names())}")
        
        data = {}
        for sheet_name in book.sheet_names():
            sheet = book[sheet_name]
            data[sheet_name] = list(sheet.rows())
            logger.info(f"  - {sheet_name}: {len(data[sheet_name])} rows")
        
        return data
        
    except ImportError:
        logger.error("pyexcel not available")
        return None
    except Exception as e:
        logger.error(f"pyexcel failed: {e}")
        return None

def try_xlrd():
    """Try loading with xlrd if available"""
    try:
        import xlrd
        import pandas as pd
        
        expected_file = "/Applications/My Apps/Bid Opt App Aug 17, 2025/PRD/Portfilio Optimizer/Excel Examples/Output Bulk Example for all 3 opt.xlsx"
        
        logger.info("Trying xlrd approach...")
        
        # Try with xlrd engine
        data = pd.read_excel(expected_file, sheet_name=None, engine='xlrd')
        logger.info(f"✅ File loaded successfully with xlrd engine")
        logger.info(f"Sheets: {list(data.keys())}")
        
        for sheet_name, df in data.items():
            logger.info(f"  - {sheet_name}: {df.shape[0]} rows × {df.shape[1]} columns")
        
        return data
        
    except ImportError:
        logger.error("xlrd not available")
        return None
    except Exception as e:
        logger.error(f"xlrd failed: {e}")
        return None

def convert_to_clean_excel():
    """Try to create a clean version of the expected file"""
    try:
        import openpyxl
        import pandas as pd
        
        expected_file = "/Applications/My Apps/Bid Opt App Aug 17, 2025/PRD/Portfilio Optimizer/Excel Examples/Output Bulk Example for all 3 opt.xlsx"
        clean_file = "/Applications/My Apps/Bid Opt App Aug 17, 2025/clean_expected_output.xlsx"
        
        logger.info("Trying to create clean version of expected file...")
        
        # Load with openpyxl and save clean version
        wb = openpyxl.load_workbook(expected_file, data_only=True)
        
        # Create new workbook
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            new_ws = new_wb.create_sheet(sheet_name)
            
            # Copy values only (no formatting)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        
        # Save clean version
        new_wb.save(clean_file)
        wb.close()
        new_wb.close()
        
        logger.info(f"✅ Clean file created: {clean_file}")
        
        # Try to load clean version with pandas
        clean_data = pd.read_excel(clean_file, sheet_name=None, na_filter=False)
        logger.info(f"✅ Clean file loads successfully with pandas")
        
        return clean_data, clean_file
        
    except Exception as e:
        logger.error(f"Clean file creation failed: {e}")
        return None, None

def main():
    logger.info("=== Alternative Excel Loading Methods Test ===")
    
    methods = [
        ("openpyxl direct", try_openpyxl_direct),
        ("xlwings", try_xlwings),
        ("pyexcel", try_pyexcel),
        ("xlrd", try_xlrd),
    ]
    
    successful_data = None
    successful_method = None
    
    # Try each method
    for method_name, method_func in methods:
        logger.info(f"\n--- Testing {method_name} ---")
        try:
            result = method_func()
            if result is not None:
                logger.info(f"✅ SUCCESS with {method_name}")
                successful_data = result
                successful_method = method_name
                break
        except Exception as e:
            logger.error(f"❌ {method_name} failed: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
    
    # If all methods failed, try creating clean file
    if successful_data is None:
        logger.info(f"\n--- All direct methods failed, trying clean file creation ---")
        clean_data, clean_file = convert_to_clean_excel()
        if clean_data is not None:
            successful_data = clean_data
            successful_method = f"clean file conversion ({clean_file})"
    
    # Report results
    if successful_data is not None:
        logger.info(f"\n🎉 SUCCESS: Loaded expected file with {successful_method}")
        if isinstance(successful_data, dict):
            for sheet_name, data in successful_data.items():
                if hasattr(data, 'shape'):  # pandas DataFrame
                    logger.info(f"  - {sheet_name}: {data.shape[0]} rows × {data.shape[1]} columns")
                elif isinstance(data, list):  # list of rows
                    logger.info(f"  - {sheet_name}: {len(data)} rows")
        return True, successful_data, successful_method
    else:
        logger.error("❌ All methods failed to load expected file")
        return False, None, None

if __name__ == "__main__":
    success, data, method = main()
    sys.exit(0 if success else 1)