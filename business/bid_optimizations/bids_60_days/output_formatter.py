"""Output formatting for Bids 60 Days optimization."""

import pandas as pd
from typing import Dict, List, Optional
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
from .constants import HELPER_COLUMNS
from business.processors.excel_base_formatter import ExcelBaseFormatter


class Bids60DaysOutputFormatter(ExcelBaseFormatter):
    """Formats output for Bids 60 Days optimization."""
    
    def __init__(self):
        super().__init__()
        self.original_columns = None
        self.blue_fill = PatternFill(
            start_color="00B0F0",
            end_color="00B0F0", 
            fill_type="solid"
        )
    
    def format_output(self, sheets_data: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """
        Format output DataFrames for Excel export.
        
        Args:
            sheets_data: Dictionary with sheet names and DataFrames
            
        Returns:
            Formatted dictionary of DataFrames
        """
        formatted_sheets = {}
        
        for sheet_name, df in sheets_data.items():
            if sheet_name == 'Targeting':
                formatted_sheets[sheet_name] = self._format_targeting_sheet(df)
            elif sheet_name == 'Bidding Adjustment':
                formatted_sheets[sheet_name] = self._format_bidding_sheet(df)
            elif sheet_name == 'For Harvesting':
                formatted_sheets[sheet_name] = self._format_harvesting_sheet(df)
            else:
                formatted_sheets[sheet_name] = df
        
        return formatted_sheets
    
    def _format_targeting_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format Targeting sheet with helper columns positioned after Bid column."""
        if df.empty:
            return df
        
        # Store original column order (first 48 columns)
        self.original_columns = [col for col in df.columns if col not in HELPER_COLUMNS and not col.startswith('_')][:48]
        
        # Ensure all helper columns exist
        for col in HELPER_COLUMNS:
            if col not in df.columns:
                df[col] = None
        
        # Use ExcelBaseFormatter method to position helpers immediately after Bid column
        df_with_correct_order = self.arrange_columns_with_helpers(
            df, self.original_columns, HELPER_COLUMNS
        )
        
        return df_with_correct_order
    
    def _format_bidding_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format Bidding Adjustment sheet (48 original columns only)."""
        if df.empty:
            return df
        
        # Use stored original columns or get first 48 non-helper columns
        if self.original_columns:
            columns = self.original_columns
        else:
            columns = [col for col in df.columns if col not in HELPER_COLUMNS and not col.startswith('_')][:48]
        
        # Keep only original columns
        columns_to_keep = [col for col in columns if col in df.columns]
        
        return df[columns_to_keep].copy()
    
    def _format_harvesting_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format For Harvesting sheet (same structure as Targeting)."""
        return self._format_targeting_sheet(df)
    
    def apply_formatting(self, workbook: Workbook, sheets_data: Dict[str, pd.DataFrame]) -> None:
        """
        Apply Excel formatting to the workbook.
        
        Args:
            workbook: Openpyxl workbook object
            sheets_data: Dictionary with sheet names and DataFrames
        """
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Apply blue headers for columns participating in processing
            if sheet_name in ['Targeting', 'For Harvesting']:
                self._apply_blue_headers(worksheet)
            
            # Pink highlighting is handled by excel_writer.py based on _needs_highlight column
    
    def _apply_blue_headers(self, worksheet) -> None:
        """Apply blue color to headers of columns participating in processing."""
        # Columns that participate in bid calculation
        participating_columns = [
            'Units', 'Bid', 'Old Bid', 'Campaign Name (Informational only)',
            'Campaign ID', 'Match Type', 'Product Targeting Expression',
            'Percentage', 'Conversion Rate', 'State'
        ] + HELPER_COLUMNS
        
        # Get header row
        header_row = worksheet[1]
        
        # Apply blue to participating columns
        for cell in header_row:
            if cell.value in participating_columns:
                cell.fill = self.blue_fill
                cell.font = Font(bold=True, color="FFFFFF")
    
    def get_column_order(self, sheet_name: str) -> List[str]:
        """
        Get the correct column order for a sheet.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            List of column names in correct order
        """
        if sheet_name == 'Bidding Adjustment':
            return self.original_columns or []
        else:  # Targeting or For Harvesting
            return (self.original_columns or []) + HELPER_COLUMNS
