"""Base Excel formatter class - DO NOT MODIFY THIS FILE.

This class provides all Excel formatting functionality that should be
consistent across ALL optimizations. New optimizations should inherit
from this class and NOT override the formatting methods.
"""

import pandas as pd
from typing import Dict, List, Optional, Any
from io import BytesIO
import logging
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelBaseFormatter:
    """
    Base class for all Excel formatters.

    Provides consistent formatting across all optimizations:
    - Text format for ID columns (prevents scientific notation)
    - Uniform column widths
    - Center alignment for all cells
    - Helper columns positioned before Bid
    - Standard colors and formatting

    DO NOT OVERRIDE THESE METHODS IN CHILD CLASSES!
    """

    # =========================================================================
    # GLOBAL CONSTANTS - NEVER CHANGE THESE
    # =========================================================================

    # Standard column width for ALL columns
    STANDARD_COLUMN_WIDTH = 15

    # Standard row height
    STANDARD_ROW_HEIGHT = 15

    # The Bid column - helpers go BEFORE this
    BID_COLUMN_NAME = "Bid"

    # Columns that MUST be formatted as text
    TEXT_FORMAT_COLUMNS = [
        "Campaign ID",
        "Ad Group ID",
        "Portfolio ID",
        "Keyword ID",
        "Product Targeting ID",
        "Campaign Id",
        "Ad Group Id",
        "Portfolio Id",
        "Keyword Id",
        "Product Targeting Id",
        "ASIN",
        "SKU",
        "Item ID",
        "Order ID",
        "Customer ID",
        "Transaction ID",
    ]

    # Standard colors
    COLORS = {"pink_error": "FFE4E1", "blue_header": "87CEEB", "gray_header": "E0E0E0"}

    # Standard number formats
    NUMBER_FORMATS = {
        # IDs - Text format
        "Campaign ID": "@",
        "Ad Group ID": "@",
        "Keyword ID": "@",
        "Product Targeting ID": "@",
        # Money
        "Bid": "0.000",
        "Old Bid": "0.000",
        "Temp Bid": "0.000",
        "Max_Bid": "0.000",
        "Base Bid": "0.000",
        "Target CPA": "0.00",
        "Adj. CPA": "0.00",
        "Max BA": "0.00",
        "Spend": "$#,##0.00",
        "Sales": "$#,##0.00",
        "CPC": "$0.000",
        # Percentages
        "Conversion Rate": "0.00%",
        "ACOS": "0.00%",
        "Percentage": "0.00",
        # Integers
        "Clicks": "#,##0",
        "Units": "#,##0",
        "Impressions": "#,##0",
        "Orders": "#,##0",
        # Decimals
        "ROAS": "0.00",
        "calc1": "0.000",
        "calc2": "0.000",
        "calc3": "0.000",
    }

    def __init__(self):
        """Initialize base formatter."""
        self.logger = logging.getLogger(self.__class__.__name__)

    # =========================================================================
    # CORE FORMATTING METHODS - DO NOT OVERRIDE
    # =========================================================================

    def prepare_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Prepare DataFrame for Excel writing.
        Converts ID columns to text to prevent scientific notation.

        ALWAYS call this before writing to Excel!

        Args:
            df: DataFrame to prepare

        Returns:
            Prepared DataFrame
        """
        df_copy = df.copy()

        for col in df_copy.columns:
            if self._should_format_as_text(col):
                # Convert to string, preserving all digits
                mask = df_copy[col].notna()
                df_copy.loc[mask, col] = df_copy.loc[mask, col].apply(
                    lambda x: str(int(x))
                    if isinstance(x, (int, float)) and x == int(x)
                    else str(x)
                    if pd.notna(x)
                    else ""
                )

        return df_copy

    def arrange_columns_with_helpers(
        self, df: pd.DataFrame, original_columns: List[str], helper_columns: List[str]
    ) -> pd.DataFrame:
        """
        Arrange columns with helpers positioned BEFORE Bid column.

        Args:
            df: DataFrame to arrange
            original_columns: Original column names
            helper_columns: Helper column names

        Returns:
            DataFrame with correct column order
        """
        # Find Bid column position
        bid_position = self._get_bid_column_position(original_columns)

        # Split columns
        before_bid = original_columns[:bid_position]
        from_bid_onwards = original_columns[bid_position:]

        # Build final order: [original...] + [helpers] + [Bid onwards...]
        final_order = before_bid + helper_columns + from_bid_onwards

        # Ensure all columns exist
        for col in final_order:
            if col not in df.columns:
                df[col] = ""

        return df[final_order]

    def write_to_excel_with_formatting(
        self,
        sheets_dict: Dict[str, pd.DataFrame],
        output_path: BytesIO = None,
        highlight_errors: bool = True,
        highlight_headers: bool = True,
        blue_header_columns: List[str] = None,
    ) -> BytesIO:
        """
        Write DataFrames to Excel with consistent formatting.

        Args:
            sheets_dict: Dictionary of sheet_name: DataFrame
            output_path: Optional BytesIO object
            highlight_errors: Whether to highlight error rows in pink
            highlight_headers: Whether to color headers
            blue_header_columns: Columns that should have blue headers

        Returns:
            BytesIO object with Excel file
        """
        if output_path is None:
            output_path = BytesIO()

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets_dict.items():
                # Prepare DataFrame
                df_prepared = self.prepare_dataframe_for_excel(df)

                # Remove internal columns
                df_clean = self._remove_internal_columns(df_prepared)

                # Write to Excel
                safe_sheet_name = sheet_name[:31]  # Excel limit
                df_clean.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                # Get worksheet for formatting
                worksheet = writer.sheets[safe_sheet_name]

                # Apply all formatting
                self._apply_all_formatting(
                    worksheet,
                    df,  # Use original df for checking highlights
                    highlight_errors,
                    highlight_headers,
                    blue_header_columns,
                )

        output_path.seek(0)
        return output_path

    def _apply_all_formatting(
        self,
        worksheet,
        df: pd.DataFrame,
        highlight_errors: bool,
        highlight_headers: bool,
        blue_header_columns: List[str] = None,
    ):
        """Apply all formatting to worksheet."""

        # 1. Apply uniform column widths
        self._apply_uniform_widths(worksheet, len(df.columns))

        # 2. Center all cells
        self._apply_center_alignment(worksheet, len(df))

        # 3. Apply header formatting
        if highlight_headers:
            self._apply_header_formatting(worksheet, df.columns, blue_header_columns)

        # 4. Apply row highlighting
        if highlight_errors:
            self._apply_row_highlighting(worksheet, df)

        # 5. Apply number formats
        self._apply_number_formats(worksheet, df.columns)

        # 6. Freeze header row
        worksheet.freeze_panes = "A2"

    def _apply_uniform_widths(self, worksheet, column_count: int):
        """Apply uniform width to ALL columns."""
        for col_idx in range(1, column_count + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = self.STANDARD_COLUMN_WIDTH

    def _apply_center_alignment(self, worksheet, row_count: int):
        """Center align all cells - OPTIMIZED VERSION."""
        center_align = Alignment(horizontal="center", vertical="center")

        # Limit alignment to first 100 rows for performance
        max_rows = min(row_count + 2, 100)

        # Apply to limited rows
        for row in range(1, max_rows):
            for col in range(1, min(worksheet.max_column + 1, 50)):  # Limit columns too
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = center_align

    def _apply_header_formatting(
        self, worksheet, columns: List[str], blue_columns: List[str] = None
    ):
        """Apply header colors and formatting."""
        blue_fill = PatternFill(
            start_color=self.COLORS["blue_header"],
            end_color=self.COLORS["blue_header"],
            fill_type="solid",
        )
        gray_fill = PatternFill(
            start_color=self.COLORS["gray_header"],
            end_color=self.COLORS["gray_header"],
            fill_type="solid",
        )

        for col_idx, col_name in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)

            # Apply color based on column type
            if blue_columns and col_name in blue_columns:
                cell.fill = blue_fill
            else:
                cell.fill = gray_fill

            # Bold font for headers
            cell.font = Font(bold=True)

    def _apply_row_highlighting(self, worksheet, df: pd.DataFrame):
        """Apply pink highlighting to error rows."""
        pink_fill = PatternFill(
            start_color=self.COLORS["pink_error"],
            end_color=self.COLORS["pink_error"],
            fill_type="solid",
        )

        # Check for highlight marker column
        if "_needs_highlight" in df.columns:
            for row_idx, needs_highlight in enumerate(df["_needs_highlight"], 2):
                if needs_highlight:
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = pink_fill

    def _apply_number_formats(self, worksheet, columns: List[str]):
        """Apply number formats to columns - OPTIMIZED VERSION."""
        # Skip number formatting if there are too many rows
        if worksheet.max_row > 1000:
            return  # Skip formatting for large files

        for col_idx, col_name in enumerate(columns, 1):
            if col_name in self.NUMBER_FORMATS:
                format_str = self.NUMBER_FORMATS[col_name]
                col_letter = get_column_letter(col_idx)

                # Apply to all rows except header
                for row in range(
                    2, min(worksheet.max_row + 1, 1000)
                ):  # Limit to 1000 rows
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = format_str

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    def _should_format_as_text(self, column_name: str) -> bool:
        """Check if column should be text formatted."""
        if column_name in self.TEXT_FORMAT_COLUMNS:
            return True

        # Check patterns
        patterns = ["ID", "Id", "ASIN", "SKU", "Code", "Number"]
        return any(pattern in column_name for pattern in patterns)

    def _get_bid_column_position(self, columns: List[str]) -> int:
        """Get position where Bid column is located."""
        try:
            return columns.index(self.BID_COLUMN_NAME)
        except ValueError:
            # Try alternatives
            for alt in ["Max CPC", "Max Bid", "Default Bid"]:
                try:
                    return columns.index(alt)
                except ValueError:
                    continue
            # Default
            return min(10, len(columns) - 1)

    def _remove_internal_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove internal tracking columns before writing."""
        internal_cols = ["_needs_highlight", "_needs_pink_highlight", "_error_type"]
        cols_to_keep = [col for col in df.columns if col not in internal_cols]
        return df[cols_to_keep]

    # =========================================================================
    # METHODS THAT CAN BE OVERRIDDEN IN CHILD CLASSES
    # =========================================================================

    def get_blue_header_columns(self) -> List[str]:
        """
        Override this in child class to specify blue header columns.

        Returns:
            List of column names that should have blue headers
        """
        return []

    def mark_rows_for_highlighting(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Override this in child class to mark rows for highlighting.
        Should add a '_needs_highlight' column with True/False values.

        Args:
            df: DataFrame to mark

        Returns:
            DataFrame with _needs_highlight column
        """
        return df
