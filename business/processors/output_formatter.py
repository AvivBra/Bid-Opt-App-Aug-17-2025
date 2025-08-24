"""Output formatter for optimization results with Bids 30 Days support."""

import pandas as pd
from typing import Dict, List, Tuple, Optional, Any
from io import BytesIO
import logging
from datetime import datetime
from config.optimization_config import (
    apply_text_format_before_write,
    apply_uniform_column_widths,
)


class OutputFormatter:
    """Formats optimization results into Excel files."""

    def __init__(self):
        """Initialize output formatter."""
        self.logger = logging.getLogger(__name__)

        # Import Bids 30 Days formatter if available
        self.bids30_formatter = None
        try:
            from business.bid_optimizations.bids_30_days.output_formatter_30_days import (
                Bids30DaysOutputFormatter,
            )

            self.bids30_formatter = Bids30DaysOutputFormatter()
        except ImportError:
            self.logger.warning("Bids 30 Days formatter not available")

    def create_output_files(
        self,
        optimization_results: Dict[str, pd.DataFrame],
        optimization_name: str = None,
    ) -> Tuple[BytesIO, BytesIO]:
        """
        Create working and clean output files.

        Args:
            optimization_results: Dictionary of DataFrames from optimization
            optimization_name: Name of the optimization (for special handling)

        Returns:
            Tuple of (working_file, clean_file) as BytesIO objects
        """

        # Check if this is Bids 30 Days optimization
        is_bids_30 = self._is_bids_30_days(optimization_results, optimization_name)

        if is_bids_30 and self.bids30_formatter:
            return self._create_bids_30_files(optimization_results)
        else:
            return self._create_standard_files(optimization_results)

    def _is_bids_30_days(
        self, results: Dict[str, pd.DataFrame], optimization_name: str = None
    ) -> bool:
        """
        Check if this is Bids 30 Days optimization.

        Args:
            results: Optimization results
            optimization_name: Optional optimization name

        Returns:
            True if Bids 30 Days optimization
        """

        # Check by name
        if optimization_name and "bids 30" in optimization_name.lower():
            return True

        # Check by presence of "For Harvesting" sheet
        if "For Harvesting" in results:
            return True

        # Check by presence of specific helper columns
        if results:
            first_sheet = next(iter(results.values()))
            bids30_columns = ["Temp Bid", "Max_Bid", "calc3"]
            if all(col in first_sheet.columns for col in bids30_columns):
                return True

        return False

    def _create_bids_30_files(
        self, optimization_results: Dict[str, pd.DataFrame]
    ) -> Tuple[BytesIO, BytesIO]:
        """
        Create output files specifically for Bids 30 Days.

        Args:
            optimization_results: Dictionary of DataFrames

        Returns:
            Tuple of (working_file, clean_file)
        """

        self.logger.info("Creating Bids 30 Days output files")

        # Get original column order (from first sheet)
        original_columns = self._get_original_columns(optimization_results)

        # Format sheets using Bids 30 formatter
        formatted_sheets = self.bids30_formatter.format_sheets(
            optimization_results, original_columns
        )

        # Create working file (with all columns and highlighting)
        working_file = self.bids30_formatter.create_excel_file(
            formatted_sheets, highlight_rows=True, highlight_headers=True
        )

        # Create clean file (minimal columns, no highlighting)
        clean_sheets = self._create_clean_sheets_bids30(formatted_sheets)
        clean_file = self.bids30_formatter.create_excel_file(
            clean_sheets, highlight_rows=False, highlight_headers=False
        )

        # Log statistics
        stats = self.bids30_formatter.get_summary_stats(formatted_sheets)
        self.logger.info(f"Bids 30 Days output created: {stats}")

        return working_file, clean_file

    def _create_clean_sheets_bids30(
        self, sheets: Dict[str, pd.DataFrame]
    ) -> Dict[str, pd.DataFrame]:
        """
        Create clean sheets for Bids 30 Days (minimal columns).

        Args:
            sheets: Formatted sheets dictionary

        Returns:
            Clean sheets dictionary
        """

        clean_sheets = {}

        # Columns to keep in clean file
        essential_columns = [
            "Entity",
            "Campaign ID",
            "Ad Group ID",
            "Portfolio Name (Informational only)",
            "Campaign Name (Informational only)",
            "Ad Group Name",
            "Keyword ID",
            "Keyword Text",
            "Match Type",
            "Product Targeting ID",
            "Product Targeting Expression",
            "Bid",
            "Operation",
        ]

        for sheet_name, df in sheets.items():
            if sheet_name == "For Harvesting":
                # For Harvesting keeps more columns for analysis
                analysis_columns = essential_columns + [
                    "Units",
                    "Clicks",
                    "Impressions",
                    "Spend",
                    "Sales",
                    "Orders",
                    "Conversion Rate",
                    "ACOS",
                    "CPC",
                ]
                columns_to_keep = [col for col in analysis_columns if col in df.columns]
            else:
                columns_to_keep = [
                    col for col in essential_columns if col in df.columns
                ]

            clean_sheets[sheet_name] = df[columns_to_keep].copy()

        return clean_sheets

    def _create_standard_files(
        self, optimization_results: Dict[str, pd.DataFrame]
    ) -> Tuple[BytesIO, BytesIO]:
        """
        Create standard output files (Zero Sales or other optimizations).

        Args:
            optimization_results: Dictionary of DataFrames

        Returns:
            Tuple of (working_file, clean_file)
        """

        self.logger.info("Creating standard output files")

        # Create working file with all data
        working_file = self._create_excel_file(
            optimization_results, include_all_columns=True
        )

        # Create clean file with essential columns only
        clean_data = self._filter_essential_columns(optimization_results)
        clean_file = self._create_excel_file(clean_data, include_all_columns=False)

        return working_file, clean_file

    def _get_original_columns(self, results: Dict[str, pd.DataFrame]) -> List[str]:
        """
        Get original column order from results.

        Args:
            results: Optimization results

        Returns:
            List of original column names
        """

        # Standard Amazon bulk file columns (48 total)
        standard_columns = [
            "Entity",
            "Operation",
            "Campaign ID",
            "Ad Group ID",
            "Portfolio Name (Informational only)",
            "Campaign Name (Informational only)",
            "Ad Group Name",
            "Keyword ID",
            "Keyword Text",
            "Match Type",
            "Campaign State (Informational only)",
            "Ad Group State (Informational only)",
            "State",
            "Daily Budget",
            "Placement",
            "Percentage",
            "Product Targeting ID",
            "Product Targeting Expression",
            "Impressions",
            "Clicks",
            "Click-through Rate",
            "Spend",
            "Sales",
            "ACOS",
            "ROAS",
            "Orders",
            "Units",
            "Conversion Rate",
            "CPC",
            "Bid",
        ]

        # Try to get actual columns from first sheet
        if results:
            first_sheet = next(iter(results.values()))
            # Filter out helper columns to get original columns
            helper_cols = [
                "Old Bid",
                "calc1",
                "calc2",
                "calc3",
                "Target CPA",
                "Base Bid",
                "Adj. CPA",
                "Max BA",
                "Temp Bid",
                "Max_Bid",
                "_needs_highlight",
            ]
            original = [col for col in first_sheet.columns if col not in helper_cols]
            if len(original) >= 40:  # Reasonable number of original columns
                return original

        return standard_columns

    def _create_excel_file(
        self, data: Dict[str, pd.DataFrame], include_all_columns: bool = True
    ) -> BytesIO:
        """
        Create Excel file from DataFrames.

        Args:
            data: Dictionary of sheet name to DataFrame
            include_all_columns: Whether to include all columns

        Returns:
            BytesIO object containing Excel file
        """

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name, df in data.items():
                # Ensure sheet name is within Excel limits
                safe_sheet_name = sheet_name[:31]

                # Apply text format to prevent scientific notation
                df_formatted = apply_text_format_before_write(df)

                # Write DataFrame to Excel
                df_formatted.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                # Apply basic formatting
                worksheet = writer.sheets[safe_sheet_name]
                self._apply_basic_formatting(worksheet, df)

                # Apply uniform column widths
                apply_uniform_column_widths(worksheet, len(df.columns))

        output.seek(0)
        return output

    def _apply_basic_formatting(self, worksheet, df: pd.DataFrame):
        """
        Apply basic formatting to worksheet.

        Args:
            worksheet: Openpyxl worksheet
            df: DataFrame with data
        """

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Header formatting
        gray_fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )

        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = gray_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # DO NOT auto-adjust column widths - uniform width already applied

        # Freeze header row
        worksheet.freeze_panes = "A2"

    def _filter_essential_columns(
        self, data: Dict[str, pd.DataFrame]
    ) -> Dict[str, pd.DataFrame]:
        """
        Filter to keep only essential columns.

        Args:
            data: Dictionary of DataFrames

        Returns:
            Filtered dictionary
        """

        essential_columns = [
            "Entity",
            "Operation",
            "Campaign ID",
            "Ad Group ID",
            "Portfolio Name (Informational only)",
            "Campaign Name (Informational only)",
            "Ad Group Name",
            "Keyword ID",
            "Keyword Text",
            "Match Type",
            "Product Targeting ID",
            "Product Targeting Expression",
            "Bid",
            "State",
        ]

        filtered_data = {}

        for sheet_name, df in data.items():
            # Keep only columns that exist
            columns_to_keep = [col for col in essential_columns if col in df.columns]
            filtered_data[sheet_name] = df[columns_to_keep].copy()

        return filtered_data

    def get_statistics(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """
        Get statistics about the output.

        Args:
            data: Dictionary of DataFrames

        Returns:
            Statistics dictionary
        """

        stats = {
            "total_sheets": len(data),
            "total_rows": sum(len(df) for df in data.values()),
            "sheets": {},
        }

        for sheet_name, df in data.items():
            stats["sheets"][sheet_name] = {"rows": len(df), "columns": len(df.columns)}

        return stats
