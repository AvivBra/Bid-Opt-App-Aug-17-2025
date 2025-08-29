"""Output formatter for Empty Portfolios optimization."""

import pandas as pd
from typing import Dict, Any
import logging
from openpyxl.styles import PatternFill
from .constants import EMPTY_PORTFOLIO_COLOR
from config.optimization_config import apply_text_format_before_write


class EmptyPortfoliosOutputFormatter:
    """Formats output for Empty Portfolios optimization."""
    
    def __init__(self):
        self.logger = logging.getLogger("empty_portfolios.output_formatter")
    
    def format_output(
        self, 
        processed_data: Dict[str, pd.DataFrame],
        writer: pd.ExcelWriter
    ) -> None:
        """
        Format and write output to Excel file.
        
        Args:
            processed_data: Dictionary with processed DataFrames
            writer: Excel writer object
        """
        try:
            # Write Sponsored Products Campaigns sheet
            campaigns_df = processed_data["Sponsored Products Campaigns"].copy()
            
            # Remove any internal columns
            if "_highlight" in campaigns_df.columns:
                campaigns_df = campaigns_df.drop("_highlight", axis=1)
            
            # Apply global text formatting to prevent scientific notation
            campaigns_df = apply_text_format_before_write(campaigns_df)
            
            campaigns_df.to_excel(
                writer, 
                sheet_name="Sponsored Products Campaigns",
                index=False
            )
            
            # Write Portfolios sheet
            portfolios_df = processed_data["Portfolios"].copy()
            
            # Track which rows need highlighting
            highlight_rows = []
            if "_highlight" in portfolios_df.columns:
                highlight_rows = portfolios_df[portfolios_df["_highlight"] == True].index.tolist()
                portfolios_df = portfolios_df.drop("_highlight", axis=1)
            
            # Apply global text formatting to prevent scientific notation
            portfolios_df = apply_text_format_before_write(portfolios_df)
            
            portfolios_df.to_excel(
                writer,
                sheet_name="Portfolios", 
                index=False
            )
            
            
            # Apply yellow highlighting to empty portfolio rows
            if highlight_rows:
                worksheet = writer.sheets["Portfolios"]
                yellow_fill = PatternFill(
                    start_color=EMPTY_PORTFOLIO_COLOR,
                    end_color=EMPTY_PORTFOLIO_COLOR,
                    fill_type="solid"
                )
                
                for row_idx in highlight_rows:
                    # Excel rows are 1-indexed, and we need to account for header
                    excel_row = row_idx + 2
                    for col_idx in range(1, len(portfolios_df.columns) + 1):
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell.fill = yellow_fill
            
            self.logger.info(f"Formatted output with {len(highlight_rows)} highlighted rows")
            
        except Exception as e:
            self.logger.error(f"Output formatting error: {str(e)}")
            raise
    
