"""Output formatter for Campaigns Without Portfolios optimization."""

import pandas as pd
from typing import Dict, List, Optional, Any
import logging
from openpyxl.styles import PatternFill
from .constants import HIGHLIGHT_COLOR, REQUIRED_COLUMNS


class CampaignsWithoutPortfoliosOutputFormatter:
    """Formats output for Campaigns Without Portfolios optimization."""
    
    def __init__(self):
        self.logger = logging.getLogger("campaigns_without_portfolios.output_formatter")
        self.updated_indices = []
    
    def format_output(
        self,
        processed_data: Dict[str, pd.DataFrame],
        highlight_rows: Optional[List[int]] = None
    ) -> pd.DataFrame:
        """
        Format the processed data for output.
        
        Args:
            processed_data: Dictionary with processed DataFrames
            highlight_rows: Optional list of row indices to highlight
            
        Returns:
            Formatted DataFrame ready for Excel output
        """
        try:
            # Get the full campaigns dataframe
            output_df = processed_data.get(
                "Sponsored Products Campaigns Full",
                processed_data.get("Sponsored Products Campaigns")
            ).copy()
            
            # Ensure all required columns are present
            for col in REQUIRED_COLUMNS:
                if col not in output_df.columns:
                    output_df[col] = ""
            
            # Reorder columns to match original order
            output_df = output_df[REQUIRED_COLUMNS]
            
            # Store indices for highlighting
            if highlight_rows:
                self.updated_indices = highlight_rows
            
            self.logger.info(f"Formatted output with {len(output_df)} rows")
            
            return output_df
            
        except Exception as e:
            self.logger.error(f"Error formatting output: {str(e)}")
            raise
    
    def apply_highlighting(self, writer, sheet_name: str, df: pd.DataFrame):
        """
        Apply yellow highlighting to updated rows.
        
        Args:
            writer: ExcelWriter object
            sheet_name: Name of the sheet
            df: DataFrame being written
        """
        if not self.updated_indices:
            return
        
        try:
            worksheet = writer.sheets[sheet_name]
            yellow_fill = PatternFill(
                start_color=HIGHLIGHT_COLOR,
                end_color=HIGHLIGHT_COLOR,
                fill_type="solid"
            )
            
            # Apply highlighting to updated rows (add 2 for header and 0-based index)
            for idx in self.updated_indices:
                excel_row = idx + 2  # +1 for header, +1 for 0-based to 1-based
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = yellow_fill
            
            self.logger.info(f"Applied highlighting to {len(self.updated_indices)} rows")
            
        except Exception as e:
            self.logger.warning(f"Could not apply highlighting: {str(e)}")
    
    def merge_with_empty_portfolios(
        self,
        campaigns_df: pd.DataFrame,
        empty_portfolios_indices: List[int]
    ) -> pd.DataFrame:
        """
        Merge results with Empty Portfolios optimization.
        
        Args:
            campaigns_df: DataFrame with campaigns without portfolios processed
            empty_portfolios_indices: Indices updated by Empty Portfolios
            
        Returns:
            Merged DataFrame
        """
        # Combine the updated indices
        all_updated_indices = list(set(self.updated_indices + empty_portfolios_indices))
        self.updated_indices = all_updated_indices
        
        self.logger.info(
            f"Merged highlighting: {len(all_updated_indices)} total rows to highlight"
        )
        
        return campaigns_df
    
    def get_summary_stats(self, processed_data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """
        Generate summary statistics for the optimization.
        
        Args:
            processed_data: Dictionary with processed DataFrames
            
        Returns:
            Dictionary with summary statistics
        """
        campaigns_df = processed_data.get("Sponsored Products Campaigns Full")
        
        stats = {
            "total_campaigns": len(campaigns_df[campaigns_df["Entity"] == "Campaign"]),
            "campaigns_updated": len(self.updated_indices),
            "target_portfolio_id": "84453417629173"
        }
        
        return stats
