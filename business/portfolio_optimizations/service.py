"""Service layer for portfolio optimizations."""

import pandas as pd
from io import BytesIO
from typing import Dict, List, Optional
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .constants import HIGHLIGHT_COLOR


class PortfolioOptimizationService:
    """Service for portfolio optimization operations."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_output_file(
        self,
        data: Dict[str, pd.DataFrame],
        updated_indices: Dict[str, List[int]]
    ) -> bytes:
        """
        Create an Excel file with highlighted changes using dedicated excel_writer.
        
        Per PRD step 45: service.py calls excel_writer.py
        
        Args:
            data: Dictionary of sheet name to DataFrame
            updated_indices: Dictionary of sheet name to list of updated row indices
            
        Returns:
            Bytes of the Excel file
        """
        self.logger.info("Creating output file using excel_writer.py per PRD step 45")
        
        # Import excel_writer as specified in PRD
        from data.writers.excel_writer import ExcelWriter
        
        # Prepare data for excel_writer
        prepared_data = {}
        for sheet_name, df in data.items():
            # Determine if this sheet was modified (has updated indices)
            was_modified = sheet_name in updated_indices and len(updated_indices[sheet_name]) > 0
            
            # Clean data before writing
            df_clean = self._prepare_dataframe(df, sheet_name, was_modified)
            
            # Add highlighting information for excel_writer
            if was_modified:
                # Mark rows that need yellow highlighting
                df_clean = self._mark_rows_for_highlighting(df_clean, updated_indices[sheet_name])
            
            prepared_data[sheet_name] = df_clean
        
        # Use dedicated excel_writer per PRD specification
        excel_writer = ExcelWriter()
        output_buffer = excel_writer.write_excel(prepared_data)
        
        # Get bytes from buffer
        result = output_buffer.getvalue()
        
        self.logger.info(f"Created output file with {len(data)} sheets using excel_writer.py")
        return result
    
    def _prepare_dataframe(self, df: pd.DataFrame, sheet_name: str, was_modified: bool) -> pd.DataFrame:
        """
        Prepare DataFrame for Excel output.
        
        Args:
            df: DataFrame to prepare
            sheet_name: Name of the sheet being prepared
            was_modified: Whether this sheet was modified by optimizations
            
        Returns:
            Cleaned DataFrame
        """
        df_clean = df.copy()
        
        if was_modified:
            # Only apply string conversion and .0 removal to modified sheets (like Portfolios)
            for col in df_clean.columns:
                # Use comprehensive ID detection for consistent formatting
                is_id_column = any(id_keyword in col for id_keyword in [
                    'Product Targeting ID', 'Campaign ID', 'Ad Group ID', 'Keyword ID', 
                    'Portfolio ID', 'ASIN', 'Target ID', 'Ad ID', 'ID'
                ])
                
                if df_clean[col].dtype in ['float64', 'int64']:
                    if is_id_column:
                        # ID columns: clean integer conversion to prevent scientific notation
                        df_clean[col] = df_clean[col].apply(
                            lambda x: str(int(x)) if pd.notna(x) and not pd.isna(x) else ''
                        )
                    else:
                        # Non-ID numeric columns: check if they should be integers (like dates and budgets)
                        is_integer_column = col in ['Budget Amount', 'Budget Start Date', 'Daily Budget', 'Start Date', 'End Date'] or 'Date' in col or 'Budget' in col
                        
                        if is_integer_column:
                            # Convert to integer first, then to string to remove .0
                            def clean_integer(x):
                                if pd.isna(x) or pd.isnull(x):
                                    return ''
                                try:
                                    return str(int(float(x)))
                                except (ValueError, TypeError):
                                    return ''
                            df_clean[col] = df_clean[col].apply(clean_integer)
                        else:
                            # Other numeric columns: convert to string and remove .0 suffix for simple integers
                            df_clean[col] = df_clean[col].apply(
                                lambda x: str(int(x)) if pd.notna(x) and not pd.isna(x) and str(x) != '' and float(x).is_integer() else (str(x) if pd.notna(x) else '')
                            )
                elif df_clean[col].dtype == 'object':
                    # Clean string columns and remove .0 suffix from numeric strings
                    df_clean[col] = df_clean[col].fillna('').astype(str)
                    if is_id_column:
                        # Extra ID column cleaning for consistent text format
                        mask = df_clean[col].str.endswith('.0') & df_clean[col].str.replace('.0', '').str.replace('-', '').str.isdigit()
                        df_clean.loc[mask, col] = df_clean.loc[mask, col].str.replace('.0', '')
                    else:
                        # Remove .0 suffix from numeric strings in non-ID columns
                        mask = df_clean[col].str.endswith('.0') & df_clean[col].str.replace('.0', '').str.replace('-', '').str.isdigit()
                        df_clean.loc[mask, col] = df_clean.loc[mask, col].str.replace('.0', '')
        else:
            # For unmodified sheets (Campaigns, Product Ad), preserve original data types
            # Only clean string columns minimally and handle floating point precision
            for col in df_clean.columns:
                if df_clean[col].dtype == 'float64':
                    # Handle ID columns vs numeric columns differently with comprehensive ID detection
                    is_id_column = any(id_keyword in col for id_keyword in [
                        'Product Targeting ID', 'Campaign ID', 'Ad Group ID', 'Keyword ID', 
                        'Portfolio ID', 'ASIN', 'Target ID', 'Ad ID', 'ID'
                    ])
                    
                    if is_id_column:
                        # ID columns should be converted to clean text format to prevent scientific notation
                        # Keep NaN as empty string, convert valid numbers to clean text
                        df_clean[col] = df_clean[col].apply(
                            lambda x: str(int(x)) if pd.notna(x) and not pd.isna(x) else ''
                        )
                    else:
                        # Non-ID numeric columns: handle integer columns vs precision columns differently
                        is_integer_column = col in ['Budget Amount', 'Budget Start Date', 'Daily Budget', 'Start Date', 'End Date'] or 'Date' in col or 'Budget' in col
                        
                        if is_integer_column:
                            # Convert integer columns to clean integer strings
                            def clean_integer(x):
                                if pd.isna(x) or pd.isnull(x):
                                    return ''
                                try:
                                    return str(int(float(x)))
                                except (ValueError, TypeError):
                                    return ''
                            df_clean[col] = df_clean[col].apply(clean_integer)
                        else:
                            # For precision columns, convert integers to clean format but preserve precision artifacts for floats
                            df_clean[col] = df_clean[col].apply(
                                lambda x: str(int(x)) if pd.notna(x) and not pd.isna(x) and str(x) != '' and float(x).is_integer() else (str(x) if pd.notna(x) else '')
                            )
                elif df_clean[col].dtype == 'object':
                    # Only handle NaN values, don't change formatting
                    df_clean[col] = df_clean[col].fillna('')
                    # Use comprehensive ID detection and clean formatting to prevent scientific notation
                    is_id_column = any(id_keyword in col for id_keyword in [
                        'Product Targeting ID', 'Campaign ID', 'Ad Group ID', 'Keyword ID', 
                        'Portfolio ID', 'ASIN', 'Target ID', 'Ad ID', 'ID'
                    ])
                    
                    if is_id_column:
                        # Clean ID columns: remove .0 suffix and ensure text format
                        mask = (df_clean[col].astype(str).str.endswith('.0') & 
                               df_clean[col].astype(str).str.replace('.0', '').str.replace('-', '').str.isdigit())
                        df_clean.loc[mask, col] = df_clean[col].astype(str).str.replace('.0', '')
                        # Force all ID values to be clean strings to prevent scientific notation
                        df_clean[col] = df_clean[col].astype(str)
        
        # Remove any columns that start with underscore (internal columns)
        cols_to_remove = [col for col in df_clean.columns if col.startswith('_')]
        if cols_to_remove:
            df_clean = df_clean.drop(columns=cols_to_remove)
        
        return df_clean
    
    def _mark_rows_for_highlighting(self, df: pd.DataFrame, updated_indices: List[int]) -> pd.DataFrame:
        """
        Mark rows that need yellow highlighting for excel_writer.
        
        Args:
            df: DataFrame to mark
            updated_indices: List of row indices that were updated
            
        Returns:
            DataFrame with highlighting markers
        """
        # Add highlighting marker column for excel_writer
        df['_needs_highlight'] = False
        
        # Mark updated rows for highlighting
        if updated_indices:
            # Ensure indices are within DataFrame bounds
            valid_indices = [idx for idx in updated_indices if 0 <= idx < len(df)]
            if valid_indices:
                df.loc[valid_indices, '_needs_highlight'] = True
                self.logger.info(f"Marked {len(valid_indices)} rows for yellow highlighting")
        
        return df
    
    def _add_highlighting(
        self,
        excel_buffer: BytesIO,
        updated_indices: Dict[str, List[int]]
    ) -> BytesIO:
        """
        Add yellow highlighting to updated rows.
        
        Args:
            excel_buffer: BytesIO containing Excel file
            updated_indices: Dictionary of sheet name to list of updated row indices
            
        Returns:
            BytesIO with highlighting added
        """
        # Load workbook
        wb = load_workbook(excel_buffer)
        
        # Create yellow fill
        yellow_fill = PatternFill(
            start_color=HIGHLIGHT_COLOR,
            end_color=HIGHLIGHT_COLOR,
            fill_type="solid"
        )
        
        # Apply highlighting to each sheet
        for sheet_name, indices in updated_indices.items():
            if not indices:
                continue
            
            # Find matching worksheet (sheet names might be truncated)
            ws = None
            for ws_name in wb.sheetnames:
                if ws_name.startswith(sheet_name[:31]):
                    ws = wb[ws_name]
                    break
            
            if ws is None:
                self.logger.warning(f"Sheet {sheet_name} not found for highlighting")
                continue
            
            # Apply highlighting
            for row_idx in indices:
                # Excel rows are 1-indexed, add 1 for header
                excel_row = row_idx + 2
                
                # Highlight entire row
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=excel_row, column=col)
                    cell.fill = yellow_fill
            
            self.logger.info(f"Highlighted {len(indices)} rows in {sheet_name}")
        
        # Save to new buffer
        output = BytesIO()
        wb.save(output)
        return output
    
    def generate_filename(self, prefix: str = "portfolio_optimized") -> str:
        """
        Generate a filename with timestamp.
        
        Args:
            prefix: Filename prefix
            
        Returns:
            Filename with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"