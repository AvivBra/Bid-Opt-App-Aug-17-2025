"""Comprehensive output file validation tests for Portfolio Optimizer."""

import pandas as pd
import numpy as np
from typing import Dict, List, Set, Any, Tuple
import logging
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


class PortfolioOutputValidator:
    """Validates portfolio optimizer output files against PRD specifications."""
    
    def __init__(self):
        self.validation_errors = []
        self.validation_warnings = []
        
    def validate_complete_output(
        self, 
        input_sheets: Dict[str, pd.DataFrame],
        output_bytes: bytes,
        updated_indices: Dict[str, List[int]],
        expected_changes: Dict[str, Dict[int, Dict[str, Any]]] = None
    ) -> Dict[str, Any]:
        """
        Run complete validation suite on output file.
        
        Args:
            input_sheets: Original input data
            output_bytes: Generated Excel file as bytes
            updated_indices: Expected updated row indices by sheet
            expected_changes: Expected cell changes by sheet->row->column->value
            
        Returns:
            Validation report with pass/fail results
        """
        logger.info("Starting comprehensive output validation")
        self.validation_errors = []
        self.validation_warnings = []
        
        # Load output Excel file
        try:
            output_sheets = self._load_excel_from_bytes(output_bytes)
        except Exception as e:
            return self._create_error_report(f"Failed to load output Excel file: {e}")
        
        # Run all validation tests
        tests = [
            ("File Structure", self._validate_file_structure, [output_sheets]),
            ("Sheet Names", self._validate_sheet_names, [output_sheets]),
            ("Row Count Preservation", self._validate_row_counts, [input_sheets, output_sheets]),
            ("Column Preservation", self._validate_column_preservation, [input_sheets, output_sheets]),
            ("Data Integrity", self._validate_data_integrity, [input_sheets, output_sheets, updated_indices]),
            ("Cell-Level Changes", self._validate_cell_changes, [input_sheets, output_sheets, expected_changes]),
            ("ID Column Formatting", self._validate_id_formatting, [output_sheets]),
            ("Yellow Highlighting", self._validate_highlighting, [output_bytes, updated_indices])
        ]
        
        results = {}
        for test_name, test_func, args in tests:
            try:
                logger.info(f"Running test: {test_name}")
                test_result = test_func(*args)
                results[test_name] = test_result
                if not test_result.get("passed", False):
                    logger.error(f"Test FAILED: {test_name}")
                else:
                    logger.info(f"Test PASSED: {test_name}")
            except Exception as e:
                logger.error(f"Test ERROR: {test_name} - {e}")
                results[test_name] = {"passed": False, "error": str(e)}
        
        # Generate final report
        return self._generate_validation_report(results)
    
    def _load_excel_from_bytes(self, excel_bytes: bytes) -> Dict[str, pd.DataFrame]:
        """Load Excel file from bytes into DataFrame dictionary."""
        buffer = BytesIO(excel_bytes)
        return pd.read_excel(buffer, sheet_name=None)
    
    def _validate_file_structure(self, output_sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Validate file has correct structure per PRD."""
        errors = []
        
        # Should have exactly 2 sheets
        if len(output_sheets) != 2:
            errors.append(f"Expected 2 sheets, found {len(output_sheets)}")
        
        # Sheets should not be empty
        for sheet_name, df in output_sheets.items():
            if df.empty:
                errors.append(f"Sheet '{sheet_name}' is empty")
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "details": f"Found {len(output_sheets)} sheets: {list(output_sheets.keys())}"
        }
    
    def _validate_sheet_names(self, output_sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Validate sheet names match PRD specifications."""
        expected_sheets = {"Sponsored Products Campaigns", "Portfolios"}
        actual_sheets = set(output_sheets.keys())
        
        errors = []
        if not expected_sheets.issubset(actual_sheets):
            missing = expected_sheets - actual_sheets
            errors.append(f"Missing required sheets: {missing}")
        
        extra = actual_sheets - expected_sheets
        if extra:
            errors.append(f"Unexpected extra sheets: {extra}")
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "details": f"Expected: {expected_sheets}, Found: {actual_sheets}"
        }
    
    def _validate_row_counts(
        self, 
        input_sheets: Dict[str, pd.DataFrame], 
        output_sheets: Dict[str, pd.DataFrame]
    ) -> Dict[str, Any]:
        """Validate no rows were lost or added."""
        errors = []
        details = {}
        
        for sheet_name in ["Sponsored Products Campaigns", "Portfolios"]:
            if sheet_name in input_sheets and sheet_name in output_sheets:
                input_count = len(input_sheets[sheet_name])
                output_count = len(output_sheets[sheet_name])
                
                details[sheet_name] = {"input": input_count, "output": output_count}
                
                if input_count != output_count:
                    errors.append(f"Row count mismatch in {sheet_name}: {input_count} -> {output_count}")
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "details": details
        }
    
    def _validate_column_preservation(
        self, 
        input_sheets: Dict[str, pd.DataFrame], 
        output_sheets: Dict[str, pd.DataFrame]
    ) -> Dict[str, Any]:
        """Validate all original columns are preserved."""
        errors = []
        details = {}
        
        for sheet_name in ["Sponsored Products Campaigns", "Portfolios"]:
            if sheet_name in input_sheets and sheet_name in output_sheets:
                input_cols = set(input_sheets[sheet_name].columns)
                output_cols = set(output_sheets[sheet_name].columns)
                
                missing_cols = input_cols - output_cols
                extra_cols = output_cols - input_cols
                
                details[sheet_name] = {
                    "input_columns": len(input_cols),
                    "output_columns": len(output_cols),
                    "missing": list(missing_cols),
                    "extra": list(extra_cols)
                }
                
                if missing_cols:
                    errors.append(f"Missing columns in {sheet_name}: {missing_cols}")
                
                # Extra columns might be OK (helper columns), just warn
                if extra_cols:
                    self.validation_warnings.append(f"Extra columns in {sheet_name}: {extra_cols}")
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "warnings": getattr(self, 'validation_warnings', []),
            "details": details
        }
    
    def _validate_data_integrity(
        self, 
        input_sheets: Dict[str, pd.DataFrame], 
        output_sheets: Dict[str, pd.DataFrame],
        updated_indices: Dict[str, List[int]]
    ) -> Dict[str, Any]:
        """Validate unchanged rows remain identical."""
        errors = []
        details = {}
        
        for sheet_name in ["Sponsored Products Campaigns", "Portfolios"]:
            if sheet_name not in input_sheets or sheet_name not in output_sheets:
                continue
                
            input_df = input_sheets[sheet_name]
            output_df = output_sheets[sheet_name]
            updated_rows = set(updated_indices.get(sheet_name, []))
            
            unchanged_rows = []
            changed_rows = []
            error_rows = []
            
            # Check each row
            min_rows = min(len(input_df), len(output_df))
            for idx in range(min_rows):
                if idx in updated_rows:
                    changed_rows.append(idx)
                    continue
                
                # Compare unchanged rows (only common columns)
                common_cols = list(set(input_df.columns) & set(output_df.columns))
                try:
                    input_row = input_df.iloc[idx][common_cols].fillna("")
                    output_row = output_df.iloc[idx][common_cols].fillna("")
                    
                    if not input_row.equals(output_row):
                        error_rows.append(idx)
                        # Find specific differences
                        differences = []
                        for col in common_cols:
                            if str(input_row[col]) != str(output_row[col]):
                                differences.append(f"{col}: '{input_row[col]}' -> '{output_row[col]}'")
                        errors.append(f"Unchanged row {idx} in {sheet_name} was modified: {differences}")
                    else:
                        unchanged_rows.append(idx)
                except Exception as e:
                    error_rows.append(idx)
                    errors.append(f"Error comparing row {idx} in {sheet_name}: {e}")
            
            details[sheet_name] = {
                "unchanged_rows": len(unchanged_rows),
                "changed_rows": len(changed_rows), 
                "error_rows": len(error_rows),
                "total_rows": min_rows
            }
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "details": details
        }
    
    def _validate_cell_changes(
        self,
        input_sheets: Dict[str, pd.DataFrame],
        output_sheets: Dict[str, pd.DataFrame], 
        expected_changes: Dict[str, Dict[int, Dict[str, Any]]] = None
    ) -> Dict[str, Any]:
        """Validate only expected cells were changed."""
        if expected_changes is None:
            return {"passed": True, "skipped": "No expected changes provided"}
        
        errors = []
        details = {}
        
        for sheet_name, sheet_changes in expected_changes.items():
            if sheet_name not in input_sheets or sheet_name not in output_sheets:
                continue
                
            input_df = input_sheets[sheet_name]
            output_df = output_sheets[sheet_name]
            
            verified_changes = 0
            missing_changes = 0
            
            for row_idx, cell_changes in sheet_changes.items():
                for col_name, expected_value in cell_changes.items():
                    try:
                        actual_value = output_df.iloc[row_idx][col_name]
                        if str(actual_value) == str(expected_value):
                            verified_changes += 1
                        else:
                            missing_changes += 1
                            errors.append(
                                f"Expected {sheet_name}[{row_idx}][{col_name}] = '{expected_value}', "
                                f"got '{actual_value}'"
                            )
                    except Exception as e:
                        missing_changes += 1
                        errors.append(f"Error validating {sheet_name}[{row_idx}][{col_name}]: {e}")
            
            details[sheet_name] = {
                "verified_changes": verified_changes,
                "missing_changes": missing_changes
            }
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "details": details
        }
    
    def _validate_id_formatting(self, output_sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Validate ID columns are formatted as text (no scientific notation)."""
        errors = []
        details = {}
        
        id_columns = [
            'Campaign ID', 'Ad Group ID', 'Keyword ID', 'Portfolio ID', 
            'Product Targeting ID', 'ASIN', 'Target ID'
        ]
        
        for sheet_name, df in output_sheets.items():
            sheet_details = {"columns_checked": 0, "scientific_notation_found": 0}
            
            for col in id_columns:
                if col in df.columns:
                    sheet_details["columns_checked"] += 1
                    
                    # Check for scientific notation (E+ or E-)
                    scientific_values = []
                    for idx, value in enumerate(df[col]):
                        if pd.notna(value):
                            str_val = str(value)
                            if 'E+' in str_val or 'E-' in str_val:
                                scientific_values.append(f"Row {idx}: {str_val}")
                                sheet_details["scientific_notation_found"] += 1
                    
                    if scientific_values:
                        errors.append(f"Scientific notation in {sheet_name}[{col}]: {scientific_values[:5]}")
            
            details[sheet_name] = sheet_details
        
        return {
            "passed": len(errors) == 0,
            "errors": errors,
            "details": details
        }
    
    def _validate_highlighting(
        self, 
        output_bytes: bytes, 
        updated_indices: Dict[str, List[int]]
    ) -> Dict[str, Any]:
        """Validate yellow highlighting on updated rows."""
        try:
            # Load workbook to check cell formatting
            buffer = BytesIO(output_bytes)
            wb = openpyxl.load_workbook(buffer)
            
            errors = []
            details = {}
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            for sheet_name, row_indices in updated_indices.items():
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                highlighted_rows = 0
                missing_highlights = 0
                
                for row_idx in row_indices:
                    excel_row = row_idx + 2  # Convert to Excel row (1-indexed + header)
                    
                    # Check if any cell in the row has yellow highlighting
                    row_has_highlight = False
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=excel_row, column=col)
                        if cell.fill and cell.fill.start_color:
                            # Check if it's yellow (FFFF00 or similar)
                            color = cell.fill.start_color.rgb
                            if color and (color == "FFFF00" or color == "00FFFF00"):
                                row_has_highlight = True
                                break
                    
                    if row_has_highlight:
                        highlighted_rows += 1
                    else:
                        missing_highlights += 1
                
                details[sheet_name] = {
                    "expected_highlights": len(row_indices),
                    "highlighted_rows": highlighted_rows,
                    "missing_highlights": missing_highlights
                }
                
                if missing_highlights > 0:
                    errors.append(f"Missing yellow highlights in {sheet_name}: {missing_highlights} rows")
            
            wb.close()
            
            return {
                "passed": len(errors) == 0,
                "errors": errors,
                "details": details
            }
            
        except Exception as e:
            return {
                "passed": False,
                "errors": [f"Could not validate highlighting: {e}"],
                "details": {"error": str(e)}
            }
    
    def _generate_validation_report(self, test_results: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        """Generate final validation report."""
        passed_tests = sum(1 for result in test_results.values() if result.get("passed", False))
        total_tests = len(test_results)
        
        all_errors = []
        all_warnings = []
        
        for test_name, result in test_results.items():
            if "errors" in result:
                all_errors.extend([f"{test_name}: {error}" for error in result["errors"]])
            if "warnings" in result:
                all_warnings.extend([f"{test_name}: {warning}" for warning in result["warnings"]])
        
        return {
            "overall_passed": passed_tests == total_tests,
            "tests_passed": passed_tests,
            "tests_total": total_tests,
            "success_rate": (passed_tests / total_tests * 100) if total_tests > 0 else 0,
            "test_results": test_results,
            "all_errors": all_errors,
            "all_warnings": all_warnings + self.validation_warnings,
            "summary": f"Validation: {passed_tests}/{total_tests} tests passed"
        }
    
    def _create_error_report(self, error_message: str) -> Dict[str, Any]:
        """Create error report for critical failures."""
        return {
            "overall_passed": False,
            "tests_passed": 0,
            "tests_total": 0,
            "success_rate": 0,
            "critical_error": error_message,
            "all_errors": [error_message],
            "summary": f"Critical Error: {error_message}"
        }


def run_output_validation_test() -> Dict[str, Any]:
    """
    Convenience function to run output validation test with portfolio optimizer.
    Returns validation report.
    """
    import pandas as pd
    from .orchestrator import PortfolioOptimizationOrchestrator
    from .service import PortfolioOptimizationService
    
    logger.info("Running comprehensive output validation test")
    
    # Create test data
    campaigns_data = pd.DataFrame({
        'Entity': ['Campaign', 'Campaign', 'Campaign'],
        'Campaign ID': ['12345678901234', '98765432109876', '55555555555555'],
        'Portfolio ID': ['999', '', '888'],
        'Operation': ['', '', '']
    })
    
    portfolios_data = pd.DataFrame({
        'Entity': ['Portfolio', 'Portfolio'],
        'Portfolio ID': ['999', '777'],
        'Portfolio Name': ['Active Portfolio', 'Empty Portfolio'],
        'Operation': ['', ''],
        'Budget Amount': ['', ''],
        'Budget Start Date': ['', '']
    })
    
    test_sheets = {
        'Sponsored Products Campaigns': campaigns_data,
        'Portfolios': portfolios_data
    }
    
    # Run optimization
    orchestrator = PortfolioOptimizationOrchestrator()
    merged_data, run_report = orchestrator.run_optimizations(
        test_sheets, ['empty_portfolios', 'campaigns_without_portfolios']
    )
    
    # Generate output file
    service = PortfolioOptimizationService()
    updated_indices = orchestrator.results_manager.get_updated_indices()
    output_bytes = service.create_output_file(merged_data, updated_indices)
    
    # Run validation
    validator = PortfolioOutputValidator()
    return validator.validate_complete_output(
        input_sheets=test_sheets,
        output_bytes=output_bytes,
        updated_indices=updated_indices
    )


if __name__ == "__main__":
    # Run test when file is executed directly
    import logging
    logging.basicConfig(level=logging.INFO)
    
    print("Running Portfolio Optimizer Output Validation Test")
    print("=" * 55)
    
    result = run_output_validation_test()
    
    print(f"\n{result['summary']}")
    print(f"Success Rate: {result['success_rate']:.1f}%")
    
    if result['all_errors']:
        print(f"\nErrors Found ({len(result['all_errors'])}):")
        for error in result['all_errors']:
            print(f"  ❌ {error}")
    
    if result['all_warnings']:
        print(f"\nWarnings ({len(result['all_warnings'])}):")
        for warning in result['all_warnings']:
            print(f"  ⚠️  {warning}")
    
    if result['overall_passed']:
        print("\n🎉 All validations passed!")
    else:
        print(f"\n💥 {result['tests_total'] - result['tests_passed']} validation(s) failed!")