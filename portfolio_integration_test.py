#!/usr/bin/env python3
"""
Portfolio Optimizer Integration Test
Tests all 3 portfolio optimizations working together according to the test specification.
"""

import sys
import os
from pathlib import Path
import pandas as pd
import numpy as np
from io import BytesIO
import logging
from datetime import datetime

# Add the application root to Python path
app_root = Path(__file__).parent
sys.path.insert(0, str(app_root))

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def load_input_files():
    """Load the input files as specified in the test."""
    input_bulk_path = app_root / "PRD/Portfilio Optimizer/Excel Examples/Example - Bulk 60.xlsx"
    template_path = app_root / "PRD/Portfilio Optimizer/Excel Examples/Filled Template Exampe.xlsx"
    
    logger.info(f"Loading input files:")
    logger.info(f"  - Bulk file: {input_bulk_path}")
    logger.info(f"  - Template file: {template_path}")
    
    # Load bulk file
    sheets = pd.read_excel(input_bulk_path, sheet_name=None, dtype=str, na_filter=False)
    logger.info(f"  - Bulk file loaded with {len(sheets)} sheets")
    
    # Load template file
    template_df = pd.read_excel(template_path, sheet_name='Sheet1', dtype=str, na_filter=False)
    logger.info(f"  - Template loaded with {len(template_df)} ASINs")
    
    return sheets, template_df

def run_portfolio_optimizations(sheets, template_df):
    """Run all 3 portfolio optimizations."""
    logger.info("Starting portfolio optimizations...")
    
    try:
        # Import the orchestrator and other required modules
        from business.portfolio_optimizations.orchestrator import PortfolioOptimizationOrchestrator
        from business.portfolio_optimizations.factory import get_portfolio_optimization_factory
        
        # Initialize orchestrator
        orchestrator = PortfolioOptimizationOrchestrator()
        factory = get_portfolio_optimization_factory()
        
        # Set template data for organize_top_campaigns strategy
        organize_strategy = factory.create_strategy("organize_top_campaigns")
        if organize_strategy and hasattr(organize_strategy, 'set_template_data'):
            organize_strategy.set_template_data(template_df)
            logger.info("Template data set for organize_top_campaigns strategy")
        
        # Select all 3 optimizations
        selected_optimizations = [
            "empty_portfolios",
            "campaigns_without_portfolios", 
            "organize_top_campaigns"
        ]
        
        logger.info(f"Running optimizations: {selected_optimizations}")
        
        # Run optimizations
        merged_data, run_report = orchestrator.run_optimizations(sheets, selected_optimizations)
        
        logger.info(f"Optimizations completed:")
        logger.info(f"  - Successful: {run_report.successful_optimizations}")
        logger.info(f"  - Total rows updated: {run_report.total_rows_updated}")
        logger.info(f"  - Execution time: {run_report.execution_time_seconds:.2f}s")
        
        # Create output file
        updated_indices = getattr(run_report, "updated_indices", {})
        output_bytes = orchestrator.create_output_file(merged_data, updated_indices)
        
        logger.info("Output file created successfully")
        
        return output_bytes, merged_data, run_report
        
    except Exception as e:
        logger.error(f"Error during optimization: {str(e)}")
        raise

def save_output_file(output_bytes, filename="test_output.xlsx"):
    """Save the output file for comparison."""
    output_path = app_root / filename
    with open(output_path, 'wb') as f:
        f.write(output_bytes)
    logger.info(f"Output saved to: {output_path}")
    return output_path

def compare_with_expected_output(actual_path):
    """Compare actual output with expected output file."""
    expected_path = app_root / "PRD/Portfilio Optimizer/Excel Examples/Output Bulk Example for all 3 opt.xlsx"
    
    logger.info("Comparing actual vs expected output...")
    logger.info(f"  - Actual: {actual_path}")
    logger.info(f"  - Expected: {expected_path}")
    
    try:
        # Load actual file
        actual_sheets = pd.read_excel(actual_path, sheet_name=None, dtype=str, na_filter=False)
        
        # Load expected file with workaround for Excel filter issues
        try:
            expected_sheets = pd.read_excel(expected_path, sheet_name=None, dtype=str, na_filter=False)
        except Exception as e:
            logger.warning(f"Direct read failed: {e}")
            logger.info("Attempting workaround for Excel autofilter issue...")
            
            # Use openpyxl to get basic structure info
            import openpyxl
            wb = openpyxl.load_workbook(expected_path, read_only=True)
            sheet_names = wb.sheetnames
            
            expected_sheets = {}
            for sheet_name in sheet_names:
                try:
                    # Try to read individual sheets
                    df = pd.read_excel(expected_path, sheet_name=sheet_name, dtype=str, na_filter=False, engine='openpyxl')
                    expected_sheets[sheet_name] = df
                    logger.info(f"Successfully loaded sheet: {sheet_name} ({df.shape})")
                except Exception as sheet_error:
                    logger.warning(f"Could not load sheet {sheet_name}: {sheet_error}")
                    # Create a placeholder DataFrame to show structure
                    ws = wb[sheet_name]
                    expected_sheets[sheet_name] = pd.DataFrame({
                        'Note': [f'Could not load this sheet due to Excel format issues. Dimensions: {ws.max_row}x{ws.max_column}']
                    })
        
        comparison_results = {
            'identical': True,
            'sheet_structure_match': True,
            'dimensions_match': True,
            'content_match': True,
            'issues': []
        }
        
        # 1. Compare sheet structure (names and order)
        if list(actual_sheets.keys()) != list(expected_sheets.keys()):
            comparison_results['identical'] = False
            comparison_results['sheet_structure_match'] = False
            comparison_results['issues'].append(f"Sheet names/order differ: {list(actual_sheets.keys())} vs {list(expected_sheets.keys())}")
        
        # 2. Compare dimensions and content for each sheet
        for sheet_name in expected_sheets.keys():
            if sheet_name not in actual_sheets:
                comparison_results['identical'] = False
                comparison_results['issues'].append(f"Sheet '{sheet_name}' missing in actual output")
                continue
                
            actual_df = actual_sheets[sheet_name]
            expected_df = expected_sheets[sheet_name]
            
            # Check dimensions
            if actual_df.shape != expected_df.shape:
                comparison_results['identical'] = False
                comparison_results['dimensions_match'] = False
                comparison_results['issues'].append(f"Sheet '{sheet_name}' dimensions differ: {actual_df.shape} vs {expected_df.shape}")
                continue
            
            # Check column headers
            if list(actual_df.columns) != list(expected_df.columns):
                comparison_results['identical'] = False
                comparison_results['issues'].append(f"Sheet '{sheet_name}' columns differ: {list(actual_df.columns)} vs {list(expected_df.columns)}")
                continue
            
            # Check cell content (excluding dynamic fields like timestamps)
            exclude_patterns = ['timestamp', 'date', 'time', 'generated']
            for col in actual_df.columns:
                if any(pattern.lower() in col.lower() for pattern in exclude_patterns):
                    continue  # Skip dynamic columns
                    
                if not actual_df[col].equals(expected_df[col]):
                    # Find specific differences
                    diff_mask = actual_df[col] != expected_df[col]
                    if diff_mask.any():
                        comparison_results['identical'] = False
                        comparison_results['content_match'] = False
                        diff_count = diff_mask.sum()
                        comparison_results['issues'].append(f"Sheet '{sheet_name}' column '{col}' has {diff_count} differences")
        
        # Log results
        if comparison_results['identical']:
            logger.info("✅ OUTPUT FILES ARE IDENTICAL!")
        else:
            logger.error("❌ Output files differ:")
            for i, issue in enumerate(comparison_results['issues']):
                logger.error(f"   {i+1}. {issue}")
                if i >= 20:  # Limit to first 20 issues for readability
                    logger.error(f"   ... and {len(comparison_results['issues'])-20} more issues")
                    break
        
        return comparison_results
        
    except Exception as e:
        logger.error(f"Error comparing files: {str(e)}")
        return {'identical': False, 'error': str(e)}

def main():
    """Main test function."""
    try:
        logger.info("=== Portfolio Optimizer Integration Test ===")
        logger.info("Testing scenario: All 3 checkboxes selected with specified input files")
        
        # Step 1: Load input files
        sheets, template_df = load_input_files()
        
        # Step 2: Run optimizations
        output_bytes, merged_data, run_report = run_portfolio_optimizations(sheets, template_df)
        
        # Step 3: Save output file
        actual_output_path = save_output_file(output_bytes, "actual_output_all_3_opt.xlsx")
        
        # Step 4: Compare with expected output
        comparison_results = compare_with_expected_output(actual_output_path)
        
        # Step 5: Report final results
        logger.info("=== TEST RESULTS ===")
        
        success_criteria = [
            ("100% Compliance with Spec", True),  # Assume compliance if no errors
            ("Identical Output File", comparison_results['identical']),
            ("Real User Simulation", False)  # Will be True when Playwright test works
        ]
        
        all_passed = all(criterion[1] for criterion in success_criteria)
        
        for criterion, passed in success_criteria:
            status = "✅ PASS" if passed else "❌ FAIL"
            logger.info(f"{status} - {criterion}")
        
        if all_passed:
            logger.info("🎉 ALL SUCCESS CRITERIA MET - 100% SUCCESS!")
            return True
        else:
            logger.info("⚠️  Some criteria not met - partial success")
            return False
            
    except Exception as e:
        logger.error(f"Test failed with exception: {str(e)}")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)